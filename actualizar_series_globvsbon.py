#!/usr/bin/env python3
"""
Completa series_globvsbon.xlsx con las ruedas que le falten.

Ese archivo es el respaldo de las series que grafica la solapa "Glob vs Bon": el canje por par,
el diferencial de TIR y el spread CCL/MEP de respaldo. El navegador las guarda en localStorage,
pero si se limpian los datos del sitio hay que volver a bajar seis años de 1816 —unos 46.000
créditos entre las tres—. Con el archivo al día, ese caso cuesta una descarga y nada más.

Corre dentro del job diario, después de actualizar_historicos.py. Mira hasta qué fecha llega cada
hoja y le pide a 1816 SÓLO los días posteriores: con el archivo al día son 22 créditos, y aunque
hayan pasado semanas sin correr, el costo escala con los días que falten y no con la historia.

Hojas (fechas en filas, tickers en columnas):
    Precio MEP   precioDirty en mep de los 10 bonos de los 5 pares
    TIR          tea en mep de los mismos (DECIMAL, tal como lo devuelve 1816)
    Precio CCL   precioDirty en ccl de AL30 y GD30

No toca historicos.xlsx ni ningún otro archivo.
"""

import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("ERROR: falta openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from precios_1816 import Cliente1816, Error1816
except ImportError:
    print("ERROR: no se pudo importar precios_1816", file=sys.stderr)
    sys.exit(1)

ARCHIVO = Path(__file__).resolve().parent / "series_globvsbon.xlsx"

# Mismos pares que PARES_CANJE en globVsBon.html. Si se agrega uno allá, va también acá.
PARES = [("AL29", "GD29"), ("AL30", "GD30"), ("AL35", "GD35"), ("AE38", "GD38"), ("AL41", "GD41")]
TICKERS = [t for par in PARES for t in par]
BENCH_CCL = ["AL30", "GD30"]

# hoja -> (campo de 1816, moneda, tickers)
HOJAS = {
    "Precio MEP": ("precioDirty", "mep", TICKERS),
    "TIR":        ("tea",         "mep", TICKERS),
    "Precio CCL": ("precioDirty", "ccl", BENCH_CCL),
}

# Fecha del canje de deuda: antes de esto los bonos no existen y 1816 tampoco tiene nada.
DESDE_CERO = "2020-09-01"


def hoy_art():
    """El runner de GitHub corre en UTC; acá interesa el día argentino."""
    return (datetime.utcnow() - timedelta(hours=3)).date()


def leer_hoja(ws):
    """-> (tickers de la cabecera, {fecha: {ticker: valor}}, última fecha o None)"""
    filas = ws.iter_rows(values_only=True)
    cab = next(filas, None)
    if not cab:
        return [], {}, None
    tickers = [str(c) for c in cab[1:] if c]
    datos, ultima = {}, None
    for fila in filas:
        if not fila or not fila[0]:
            continue
        f = str(fila[0])[:10]
        if len(f) != 10 or f[4] != "-":
            continue
        datos[f] = {tickers[i]: v for i, v in enumerate(fila[1:len(tickers) + 1]) if v not in (None, "")}
        if ultima is None or f > ultima:
            ultima = f
    return tickers, datos, ultima


def escribir_hoja(ws, tickers, datos):
    ws.delete_rows(1, ws.max_row or 1)
    ws.append(["Fecha"] + tickers)
    for f in sorted(datos):
        fila = datos[f]
        ws.append([f] + [fila.get(t, "") for t in tickers])


def main():
    if not ARCHIVO.exists():
        print(f"ERROR: no está {ARCHIVO.name}. Se genera desde la solapa con 'Descargar respaldo'.",
              file=sys.stderr)
        return 1

    try:
        cli = Cliente1816()
    except Exception as e:
        print(f"ERROR: no se pudo crear el cliente de 1816: {e}", file=sys.stderr)
        return 1

    wb = load_workbook(ARCHIVO)
    hasta = hoy_art().isoformat()
    cambios = 0

    for nombre, (campo, moneda, tickers_esperados) in HOJAS.items():
        if nombre not in wb.sheetnames:
            print(f"  {nombre}: no está en el archivo, se saltea")
            continue
        ws = wb[nombre]
        tickers, datos, ultima = leer_hoja(ws)
        if not tickers:
            tickers = tickers_esperados

        # Se reanuda DESDE la última fecha con dato, inclusive: esa rueda se vuelve a pedir por si
        # cuando se guardó todavía era provisoria. Es el mismo criterio que usa el frontend.
        desde = ultima or DESDE_CERO
        if desde > hasta:
            print(f"  {nombre}: al día ({ultima})")
            continue

        try:
            filas = cli.series(tickers, [campo], moneda=moneda, fecha_inicial=desde, fecha_final=hasta)
        except Error1816 as e:
            # No se aborta todo: puede fallar una hoja y las otras andar. El archivo queda con lo
            # que se pudo y mañana se reintenta desde donde quedó.
            print(f"  {nombre}: 1816 falló ({e}), se deja como estaba", file=sys.stderr)
            continue

        nuevos = 0
        for r in filas:
            v = r.get(campo)
            if v is None:
                continue
            f = str(r.get("fecha", ""))[:10]
            tk = r.get("ticker")
            if len(f) != 10 or tk not in tickers:
                continue
            if datos.get(f, {}).get(tk) != v:
                nuevos += 1
            datos.setdefault(f, {})[tk] = v

        if nuevos:
            escribir_hoja(ws, tickers, datos)
            cambios += nuevos
        ruedas = sorted(datos)
        print(f"  {nombre}: {nuevos} valores nuevos · {len(ruedas)} ruedas · hasta {ruedas[-1] if ruedas else '—'}")

    if cambios:
        wb.save(ARCHIVO)
        print(f"{ARCHIVO.name} actualizado ({cambios} valores)")
    else:
        print(f"{ARCHIVO.name} sin cambios")
    return 0


if __name__ == "__main__":
    sys.exit(main())
