#!/usr/bin/env python3
"""
Rellena las ruedas que quedaron incompletas en historicos.xlsx, pidiéndole los precios a 1816.

POR QUÉ EXISTE. El job diario guarda el cierre del día y sigue de largo: si 1816 contesta 429 no
hay segunda oportunidad, la rueda queda con ~70 tickers en vez de ~165 y así se queda para siempre.
Pasó cinco veces entre junio y agosto de 2026, casi siempre porque el repo del colega corría su
propio job con la misma API key a la misma hora. El cron de acá se corrió a las 17:20 para que no
se pisen, pero eso arregla las ruedas futuras, no las que ya están rotas: para eso es este script.

Lo que falta en esas ruedas son subsoberanos y ONs. No es casual: el fallback es Eco Valores, que
no los publica. Los pesos y los soberanos se salvan; el resto desaparece.

QUÉ CUESTA. Se pide una fecha por vez, no el rango completo: el costo de /series es
tickers x campos x días, así que cinco días sueltos salen ~475 créditos y el rango 15/06 a 14/08
saldría ~4.000 por exactamente el mismo dato.

USO
    python completar_ruedas.py                      # detecta las ruedas rotas y las completa
    python completar_ruedas.py 2026-08-12 ...       # sólo esas fechas
    python completar_ruedas.py --dry-run            # dice qué haría y cuánto costaría, sin pedir

Nunca pisa un valor que ya está: sólo llena celdas vacías.
"""

import sys
from statistics import median

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl", file=sys.stderr)
    sys.exit(1)

from actualizar_historicos import (
    CAMPO_1816, FACTOR_ESCALA, HISTORICOS_FILE, leer_tickers,
)

try:
    from precios_1816 import Cliente1816, Error1816
except ImportError:
    print("ERROR: no se pudo importar precios_1816", file=sys.stderr)
    sys.exit(1)

# Cuántos tickers tienen que faltar para considerar que la rueda se rompió. Una rueda sana pierde
# uno o dos por falta de operaciones; las rotas pierden noventa. Con el umbral en el medio no hace
# falta distinguir "no operó" de "falló la API": ningún día normal se acerca a este número.
UMBRAL_ROTA = 20

# Reintentos ante 429. Mismo criterio que el respaldo de Glob vs Bon: el limitador de 1816 es
# global por API key, así que si otro proceso está consumiendo hay que esperar de verdad.
ESPERAS = [15, 45, 90]


def leer_historicos(ws):
    """-> (tickers de la cabecera, {fecha: {ticker: valor}})"""
    it = ws.iter_rows(values_only=True)
    cab = next(it, None) or []
    tickers = [str(c) for c in cab[1:] if c]
    datos = {}
    for fila in it:
        if not fila or not fila[0]:
            continue
        f = str(fila[0])[:10]
        if len(f) != 10 or f[4] != "-":
            continue
        datos[f] = {tickers[i]: v for i, v in enumerate(fila[1:len(tickers) + 1])
                    if v not in (None, "")}
    return tickers, datos


def ruedas_sanas(datos, ventana=10, piso=0.9):
    """Fechas que sirven de referencia: las que tienen casi tantos tickers como la mejor rueda de
    su entorno.

    La referencia es LOCAL a propósito. El archivo fue creciendo —los subsoberanos y las ONs se
    sumaron sobre la marcha—, así que una rueda de enero con 70 tickers está completa y una de
    agosto con 70 está rota; comparar todo contra un número fijo confundiría las dos cosas.

    Y el piso es 0.9 y no algo laxo porque si no las ruedas rotas se validan entre ellas: el 09/07
    tomaba como referencia al 10/07, que también había fallado, y así se "perdían" 21 tickers que
    en realidad faltaban en las dos.
    """
    fs = sorted(datos)
    sanas = set()
    for i, f in enumerate(fs):
        entorno = fs[max(0, i - ventana):i + ventana + 1]
        if len(datos[f]) >= piso * max(len(datos[x]) for x in entorno):
            sanas.add(f)
    return sanas


def faltantes(datos, fecha, candidatos, sanas):
    """Tickers que en `fecha` están vacíos pero sí cotizan en las ruedas sanas de alrededor.

    Se exige presencia ANTES y DESPUÉS para no perseguir precios que no existen: un bono emitido
    la semana siguiente, o uno que venció, falta con todo derecho y 1816 tampoco lo tiene. Para la
    última rueda del archivo no hay "después" y alcanza con el lado izquierdo.
    """
    fs = sorted(datos)
    i = fs.index(fecha)
    izq = [f for f in fs[:i] if f in sanas]
    der = [f for f in fs[i + 1:] if f in sanas]
    if not izq:
        return []
    ctx = set(datos[izq[-1]])
    if der:
        ctx &= set(datos[der[0]])
    return sorted((ctx & candidatos) - set(datos[fecha]))


def pedir(cli, tickers, moneda, fecha):
    """-> {ticker: precio} para esa fecha, o None si 1816 no contestó.

    Se pide una ventana de dos días y se filtra, en vez de pedir la fecha sola: con
    fechaInicial == fechaFinal la API contesta 200 pero sin un solo punto. Cuesta el doble
    (el costo es tickers x campos x días) y sigue siendo mucho más barato que pedir el rango
    entero: para las seis ruedas son ~950 créditos contra ~4.000.
    """
    import time
    from datetime import date as _date, timedelta
    y, m, d = map(int, fecha.split("-"))
    hasta = (_date(y, m, d) + timedelta(days=1)).isoformat()
    ultimo = None
    for espera in [0] + ESPERAS:
        if espera:
            print(f"      reintentando en {espera}s...")
            time.sleep(espera)
        try:
            filas = cli.series(tickers, [CAMPO_1816], moneda=moneda,
                               fecha_inicial=fecha, fecha_final=hasta)
            res = {r["ticker"]: r[CAMPO_1816] for r in filas
                   if str(r.get("fecha", ""))[:10] == fecha and r.get(CAMPO_1816) is not None}
            print(f"      {len(filas)} filas, {len(res)} con precio en {fecha}")
            return res
        except Error1816 as e:
            ultimo = e
            if "429" not in str(e) and "Demasiadas" not in str(e):
                break            # no es rate limit: reintentar no ayuda
        except Exception as e:
            ultimo = e
    print(f"      1816 falló ({ultimo}), se deja como estaba", file=sys.stderr)
    return None


def main(argv):
    dry = "--dry-run" in argv
    pedidas = [a for a in argv if not a.startswith("--")]

    wb = load_workbook(HISTORICOS_FILE)
    ws = wb["Historicos"]
    tickers, datos = leer_historicos(ws)
    col = {t: i + 2 for i, t in enumerate(tickers)}
    fila_de = {str(ws.cell(row=r, column=1).value)[:10]: r for r in range(2, ws.max_row + 1)}

    # Sólo los que mapean a 1816: lo que no mapea nunca se va a poder completar por acá.
    items = [it for it in leer_tickers() if it["t1816"] and it["moneda"]]
    moneda_de = {it["eco"]: it["moneda"] for it in items}
    t1816_de = {it["eco"]: it["t1816"] for it in items}
    candidatos = set(moneda_de)

    sanas = ruedas_sanas(datos)

    if pedidas:
        objetivo = [f for f in pedidas if f in datos]
        for f in pedidas:
            if f not in datos:
                print(f"AVISO: {f} no es una rueda del archivo, se saltea", file=sys.stderr)
    else:
        objetivo = [f for f in sorted(datos)
                    if len(faltantes(datos, f, candidatos, sanas)) >= UMBRAL_ROTA]

    if not objetivo:
        print("No hay ruedas incompletas.")
        return 0

    plan = {f: faltantes(datos, f, candidatos, sanas) for f in objetivo}
    costo = sum(len(v) for v in plan.values())
    print(f"{len(objetivo)} ruedas a completar · {costo} créditos estimados")
    for f in objetivo:
        print(f"  {f}: {len(datos[f]):3d} tickers, faltan {len(plan[f]):3d}")
    if dry:
        return 0

    cli = Cliente1816()
    escritos = raros = 0

    for f in objetivo:
        print(f"\n{f}:")
        porm = {}
        for t in plan[f]:
            porm.setdefault(moneda_de[t], []).append(t)

        for moneda, lote in porm.items():
            inv = {t1816_de[t]: t for t in lote}
            print(f"   {moneda}: {len(lote)} tickers")
            res = pedir(cli, sorted(inv), moneda, f)
            if res is None:
                continue

            for t1816, precio in res.items():
                eco = inv.get(t1816)
                if not eco or not isinstance(precio, (int, float)) or precio <= 0:
                    continue
                r = fila_de.get(f)
                c = col.get(eco)
                if not r or not c or ws.cell(row=r, column=c).value not in (None, ""):
                    continue     # no se pisa nada que ya esté

                # Misma guarda de escala que el job diario: un precio que se aparta de la mediana
                # del propio ticker por más de 5x es otra moneda, no un movimiento de mercado.
                previos = [v for fx in sorted(datos) for v in [datos[fx].get(eco)]
                           if isinstance(v, (int, float)) and v > 0][-20:]
                if len(previos) >= 3:
                    m = median(previos)
                    if precio > m * FACTOR_ESCALA or precio < m / FACTOR_ESCALA:
                        print(f"      {eco}: {precio} vs mediana {m:.2f}, descartado")
                        raros += 1
                        continue

                ws.cell(row=r, column=c).value = precio
                datos[f][eco] = precio
                escritos += 1

    if escritos:
        wb.save(HISTORICOS_FILE)
        print(f"\n{HISTORICOS_FILE} actualizado: {escritos} precios"
              + (f", {raros} descartados por escala" if raros else ""))
    else:
        print("\nSin cambios.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
