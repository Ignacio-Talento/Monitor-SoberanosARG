#!/usr/bin/env python3
"""Extrae los datos duros de la rueda para el informe diario.

QUÉ HACE Y QUÉ NO. Este script NO redacta el informe: junta los números y los deja en un JSON
versionado. El análisis —qué se movió, qué vale la pena mirar, qué trade aparece— lo escribe
después Claude leyendo este archivo. La división es a propósito:

  - los números tienen que ser reproducibles y auditables, así que salen de un script determinístico
    que queda commiteado junto a su resultado;
  - el análisis cambia todas las semanas según lo que pase en el mercado, y codificar reglas del
    tipo "si el spread supera X avisá" envejece mal y termina avisando de lo que ya no importa.

Guardar el JSON además deja archivo: se puede releer el informe de cualquier rueda, y comparar el
dataset de hoy contra el de hace un mes sin volver a gastar créditos de 1816.

POR QUÉ CORRE EN GITHUB ACTIONS. La API key de 1816 vive como Secret del repo y no está en ninguna
máquina local. El proxy de producción tampoco sirve como atajo: está detrás de Cloudflare Access y
devuelve 302 al login cuando no hay sesión de navegador.

QUÉ PIDE Y CUÁNTO CUESTA. Una sola llamada a /series por lote de 10 tickers, con dos ruedas —la de
hoy y la anterior— y cuatro campos. 1816 cobra tickers x campos x días, así que son unos 1.500
créditos por corrida: contra los ~21.000 que consume un día de monitor abierto, es marginal.

Se piden DOS ruedas y no una porque el informe es sobre la variación, y la rueda anterior tiene que
venir de la misma fuente que la de hoy. Sacar el precio de ayer de historicos.xlsx y la tasa de
1816 mezclaría convenciones —el Excel guarda precio dirty en la moneda de cada hoja— y las
variaciones saldrían con ruido que no es del mercado.

TIPO DE INFORME. El script decide solo si además del diario corresponde el semanal (viernes o
última rueda de la semana) y el mensual (última rueda hábil del mes), mirando el calendario de
feriados de Argentina. No hace falta un cron por tipo: corre todos los días hábiles y avisa en
`tipos` qué cierres caen hoy.

SALIDA: informes/datos_AAAA-MM-DD.json
"""
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from statistics import median

import requests

from openpyxl import load_workbook

from actualizar_historicos import (INSTRUMENTOS_FILE, cliente_1816, hoy_art,
                                    leer_tickers)
from macro_informe import datos_macro

DIR_INFORMES = Path("informes")

# Campos que se le piden a 1816. Son los mismos cuatro que pide el monitor para los instrumentos
# sin cronograma cargado (CAMPOS_IND en functions/api/precios.js), así que el informe y la pantalla
# hablan de los mismos números.
CAMPOS = ["precioDirty", "tea", "durationMod", "paridad"]

# Hoja de Instrumentos.xlsx -> familia del informe. Son las que pidió el usuario, con el nombre con
# el que se las nombra en la mesa. LECAPS y TASA FIJA van juntas: es una sola curva en pesos a tasa
# fija, y separarlas parte el tramo corto del largo del mismo instrumento económico.
FAMILIAS = {
    "LECAPS": "LECAPs y tasa fija",
    "TASA FIJA": "LECAPs y tasa fija",
    "CER": "CER",
    "TAMAR": "TAMAR",
    "Duales": "Duales",
    "USD Linked": "Dólar linked",
    "USD Bonares": "Bonares",
    "USD Globales": "Globales",
    "USD Bopreales": "BOPREALes",
    # Las ONs NO se clasifican por hoja: ver mapa_ley_ons(). Se dejan acá para que el instrumento
    # entre al universo, pero la familia se pisa después con la ley que declara el Excel.
    "ON USD": "ONs",
    "ONs": "ONs",
    "Subsoberanos": "Subsoberanos",
}

# La métrica que se mira en cada familia además del precio. En pesos a tasa fija lo que se negocia
# es la tasa —el precio de una LECAP sube todos los días por el mero paso del tiempo, así que su
# variación de precio no dice casi nada—; en hard dollar se miran las dos, precio y rendimiento.
METRICA = {
    "LECAPs y tasa fija": "TEA",
    "CER": "TIR real",
    "TAMAR": "TEA",
    "Duales": "TEA",
    "Dólar linked": "TIR",
    "Bonares": "TIR",
    "Globales": "TIR",
    "BOPREALes": "TIR",
    "ONs ley local": "TIR",
    "ONs ley NY": "TIR",
    "ONs": "TIR",                 # sólo si alguna quedó sin ley declarada
    "Subsoberanos": "TIR",
}

FERIADOS_API = "https://api.argentinadatos.com/v1/feriados/{anio}"

# El runner de GitHub corre en UTC. Sin esto, datetime.now() daba las 17:08 para una corrida de las
# 14:08 de Buenos Aires, y el informe del 2026-08-28 salió diciendo "tomados a las 17:08 con el
# mercado ya cerrado" cuando el mercado estaba abierto y los precios eran intradía. Con el tzinfo
# puesto el offset viaja en el propio string (-03:00) y no hay forma de leerlo mal.
ART = timezone(timedelta(hours=-3))


# Tipo de pata en Instrumentos.xlsx -> sufijo del ticker en 1816. Verificado contra el catálogo el
# 2026-08-28: para TXMD8 existen "TXMD8 @CER" y "TXMD8 @TAMAR", y cada uno devuelve la tasa de SU
# pata —6,18% real y 40,71% nominal—, mientras el ticker pelado devuelve la de la pata que manda.
SUFIJO_PATA = {"CER": "@CER", "TAMAR": "@TAMAR", "LECAP": "@Tasa Fija", "LINKED": "@USD-L"}


def patas_duales():
    """[(ticker, tipo, ticker1816)] con las patas de cada dual, leídas de la hoja Duales.

    POR QUÉ SE PUEDEN PEDIR. Un dual paga el máximo entre sus dos patas, así que su TIR "entera" es
    la de la pata que domina y la otra no aparece por ningún lado: con un solo número no se puede
    ubicar el instrumento ni en la curva CER ni en la TAMAR. 1816 publica las dos por separado, con
    un ticker por pata, y eso es exactamente lo que hace falta para dibujarlas.

    LA PATA DOMINADA DEVUELVE NULL, y no es un error sino información: 1816 sólo calcula
    indicadores para la pata que va a pagar. "TTS26 @Tasa Fija" viene con todo en null porque hoy
    manda la TAMAR. Así se identifica la pata in the money sin calcular nada.

    El tipo sale del Excel y no del catálogo de 1816: la hoja ya tiene una fila por pata con su
    tipo, es gratis y no gasta una consulta por dual.
    """
    wb = load_workbook(INSTRUMENTOS_FILE, data_only=True)
    if "Duales" not in wb.sheetnames:
        return []
    filas = list(wb["Duales"].iter_rows(values_only=True))
    out, vistos = [], set()
    for f in filas:
        if not f or not f[0] or str(f[0]).strip() in ("Ticker", "None"):
            continue
        tk, tipo = str(f[0]).strip(), str(f[1] or "").strip().upper()
        suf = SUFIJO_PATA.get(tipo)
        if not suf or (tk, tipo) in vistos:
            continue
        vistos.add((tk, tipo))
        out.append((tk, tipo, f"{tk} {suf}"))
    return out


def mapa_ons():
    """ticker -> {"ley": "local"|"ny", "moneda": "mep"|"ccl"}, leído de la hoja ONs.

    HACE FALTA PORQUE LA HOJA NO ES LA LEY. Al armar el informe se supuso que la hoja "ON USD" eran
    las ONs de ley local y la hoja "ONs" las de ley NY, y las dos cosas eran falsas:

      · "ON USD" no es un listado de instrumentos sino una tabla de CRONOGRAMA DE FLUJOS —columnas
        Fecha, Valor Residual, Cupón, Renta, Amortización—. Sus 56 filas son cupones de apenas
        cuatro tickers, que además ya figuran en la otra hoja. El informe del 28/08/2026 reportó
        por eso una familia "ONs ley local" de cuatro instrumentos.
      · "ONs" tiene las dos leyes juntas y las distingue en su columna Ley: 61 local y 36 ny. Todas
        salieron reportadas como ley NY.

    Con la ley leída del Excel, un instrumento va a su familia venga de la hoja que venga.

    La moneda se lee acá SÓLO para poder controlarla contra la que ya resolvió leer_tickers() y para
    informarla en el resumen. La regla es la misma que aplica el monitor en monedaDeON() (ons.html)
    y que actualizar_historicos.py ya implementa en moneda_on(): las de ley NY van todas al CCL, y
    las locales en la moneda en la que PAGAN, que es lo que dice la columna Divisa, scrapeada del
    domicilio de pago. Hoy son 43 al CCL —36 de ley NY y 7 locales que pagan en cable— y 54 al MEP.
    """
    wb = load_workbook(INSTRUMENTOS_FILE, data_only=True)
    if "ONs" not in wb.sheetnames:
        return {}
    ws = wb["ONs"]
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {}
    cab = [str(c).strip().lower() if c else "" for c in filas[0]]
    if "ticker" not in cab or "ley" not in cab:
        print("AVISO: la hoja ONs no tiene columnas Ticker y Ley; no se separan por legislación")
        return {}
    i_tk, i_ley = cab.index("ticker"), cab.index("ley")
    i_div = cab.index("divisa") if "divisa" in cab else None
    out = {}
    for f in filas[1:]:
        if not f or not f[i_tk] or not f[i_ley]:
            continue
        ley = str(f[i_ley]).strip().lower()
        div = str(f[i_div] or "").strip().upper() if i_div is not None else ""
        out[str(f[i_tk]).strip()] = {
            "ley": ley,
            "moneda": "ccl" if (ley == "ny" or div == "CCL") else "mep",
        }
    return out


def feriados(anio):
    """Feriados nacionales del año, como set de date. Si la API no responde, se sigue sin ellos.

    Sólo afecta a la detección de cierre de mes y de semana: sin feriados, el informe mensual podría
    salir un día tarde. Se avisa en el JSON en vez de abortar la corrida entera por eso.
    """
    try:
        r = requests.get(FERIADOS_API.format(anio=anio), timeout=15)
        r.raise_for_status()
        return {date.fromisoformat(f["fecha"]) for f in r.json()}
    except Exception as e:                                    # noqa: BLE001
        print(f"AVISO: no se pudieron leer los feriados de {anio} ({e})")
        return None


def es_habil(d, fer):
    return d.weekday() < 5 and (fer is None or d not in fer)


def proxima_habil(d, fer):
    x = d + timedelta(days=1)
    while not es_habil(x, fer):
        x += timedelta(days=1)
    return x


def tipos_de_cierre(d, fer):
    """Qué cierres caen en esta rueda: siempre 'diario', más 'semanal' y/o 'mensual'.

    Se define por la PRÓXIMA rueda hábil y no por el día de la semana o el número del día: un
    viernes feriado no es cierre semanal, y el 31 puede caer domingo. Mirando hacia adelante, la
    última rueda de la semana es aquella cuya siguiente hábil ya cayó en otra semana, y la última
    del mes, aquella cuya siguiente hábil ya cambió de mes.
    """
    tipos = ["diario"]
    sig = proxima_habil(d, fer)
    if sig.isocalendar()[:2] != d.isocalendar()[:2]:
        tipos.append("semanal")
    if (sig.year, sig.month) != (d.year, d.month):
        tipos.append("mensual")
    return tipos


def ultima_rueda_de_periodo_anterior(d, fer, periodo):
    """Última rueda hábil de la semana o del mes ANTERIOR a la de d.

    Es contra esta rueda que se mide el cierre semanal y el mensual. Se busca hacia atrás desde el
    primer día del período de d, saltando fines de semana y feriados: para el viernes 28/08/2026 la
    referencia semanal es el viernes 21 y la mensual, el jueves 31/07.
    """
    if periodo == "semanal":
        x = d - timedelta(days=d.weekday() + 1)          # domingo anterior al lunes de esta semana
    else:
        x = d.replace(day=1) - timedelta(days=1)         # último día del mes anterior
    while not es_habil(x, fer):
        x -= timedelta(days=1)
    return x


def rueda_anterior_habil(d, fer):
    x = d - timedelta(days=1)
    while not es_habil(x, fer):
        x -= timedelta(days=1)
    return x


def pedir_series(cli, items, desde, hasta):
    """Indicadores de cada instrumento en el rango, agrupados por moneda.

    Se agrupa por moneda porque 1816 la toma como parámetro de la request, no del ticker: pedir un
    global en 'ars' devuelve el precio en pesos y arruinaría la comparación contra su propia serie.
    """
    por_moneda = defaultdict(list)
    for it in items:
        if it["t1816"] and it["moneda"]:
            por_moneda[it["moneda"]].append(it["t1816"])

    datos = defaultdict(dict)          # ticker1816 -> fecha -> {campo: valor}
    for moneda, tickers in por_moneda.items():
        for i in range(0, len(tickers), 10):        # /series topea en 10 tickers por request
            lote = tickers[i:i + 10]
            try:
                filas = cli.series(lote, CAMPOS, moneda=moneda,
                                   fecha_inicial=desde, fecha_final=hasta)
            except Exception as e:                            # noqa: BLE001
                print(f"AVISO: lote {lote} en {moneda} falló ({e})")
                continue
            for f in filas:
                tk, fe = f.get("ticker"), f.get("fecha")
                if tk and fe:
                    datos[tk][fe] = {c: f.get(c) for c in CAMPOS}
    return datos


def pct(x):
    """1816 devuelve tea y paridad como FRACCIÓN: 0,2672 es una TEA de 26,72%.

    Verificado el 2026-08-28 contra la pantalla: S30S6 vino con tea 0,26726 y la solapa Sintéticos
    mostraba 26,68% para el mismo instrumento. Sin esta conversión el informe reportaba tasas cien
    veces más chicas, y peor: el redondeo a tres decimales dejaba una TEA del 26,7% sin el decimal
    que importa, y una variación de -1,5 pp aparecía como "-0,015 pp", o sea como si no se hubiera
    movido nada.

    durationMod NO se toca: viene en años y así se usa.
    """
    return None if x is None else x * 100


def variacion(hoy, ayer):
    if hoy is None or ayer in (None, 0):
        return None
    return round((hoy / ayer - 1) * 100, 3)


def delta(hoy, ayer):
    if hoy is None or ayer is None:
        return None
    return round(hoy - ayer, 3)


def redondear(x, n=3):
    return None if x is None else round(x, n)


def resumir(valores):
    """Mediana, mínimo y máximo de una lista que puede venir con huecos."""
    v = [x for x in valores if x is not None]
    if not v:
        return None
    return {"mediana": round(median(v), 3), "min": round(min(v), 3),
            "max": round(max(v), 3), "n": len(v)}


def main():
    cli = cliente_1816()
    if cli is None:
        print("ERROR: sin cliente de 1816. Hace falta el secret API_1816_KEY.")
        return 1

    hoy = hoy_art()          # ya viene como date, en calendario argentino
    fer = feriados(hoy.year)
    if not es_habil(hoy, fer):
        print(f"{hoy} no es rueda hábil; no se arma informe.")
        return 0

    ayer = rueda_anterior_habil(hoy, fer)
    tipos = tipos_de_cierre(hoy, fer)
    items = leer_tickers()

    # CONTROL DE QUE LA MONEDA COINCIDE CON LA DEL MONITOR. No corrige nada: leer_tickers() ya
    # resuelve bien la punta de cada ON —moneda_on() en actualizar_historicos.py aplica la misma
    # regla que monedaDeON() en ons.html: ley NY al CCL, locales según su columna Divisa—. Lo que
    # hace es avisar si alguna vez divergen.
    #
    # Vale la pena tenerlo porque son dos implementaciones de la misma regla en dos lenguajes, y
    # nada obliga a que se muevan juntas: el día que alguien cambie una y no la otra, el informe
    # va a estar mostrando una punta y la pantalla otra, y sin este control no se enteraría nadie.
    ons = mapa_ons()
    discrepan = [it["eco"] for it in items
                 if ons.get(it["eco"]) and it["moneda"] != ons[it["eco"]]["moneda"]]
    print(f"Universo: {len(items)} instrumentos · ruedas {ayer} y {hoy}")
    if discrepan:
        print(f"ATENCIÓN: {len(discrepan)} ONs con moneda distinta a la del monitor: "
              f"{', '.join(discrepan[:12])}")

    datos = pedir_series(cli, items, ayer.isoformat(), hoy.isoformat())
    print(f"1816 devolvió datos de {len(datos)} tickers")

    # SEGUNDA PASADA EN MEP PARA LO QUE SE VALÚA EN CCL.
    #
    # La tabla por familia muestra cada grupo en la punta del monitor, y ahí globales, subsoberanos
    # y ONs de ley NY van al CCL. Pero comparar un global contra su bonar gemelo con esas puntas
    # cruzadas da un número que sale, es plausible, y no significa lo que dice: la propia solapa
    # Glob vs Bon lo tiene documentado —el canje de AL29/GD29 pasa de +4,02% a +0,17% según se
    # mezclen o no—. Por eso esa solapa descarta lo que no esté en MEP en vez de convertirlo.
    #
    # Acá se hace lo mismo pero sin perder el par: se pide una segunda vez en MEP sólo lo que se
    # valúa en CCL, y el instrumento queda con las dos puntas. La tabla sigue usando la del
    # monitor; cualquier comparación entre familias usa la homogénea.
    #
    # Cuesta unos 430 créditos —54 tickers por 4 campos por 2 ruedas— contra los ~1.500 del pedido
    # principal. Convertir con el canje habría salido gratis, pero mete un supuesto propio en cada
    # número; pedirlo es exacto y barato.
    # PATAS DE LOS DUALES. Son pocas —dos por dual— así que el costo es marginal: catorce tickers
    # por cuatro campos por dos ruedas, unos 110 créditos.
    patas = patas_duales()
    datos_patas = {}
    if patas:
        falsos = [{"eco": t1816, "t1816": t1816, "moneda": "ars"} for _, _, t1816 in patas]
        datos_patas = pedir_series(cli, falsos, ayer.isoformat(), hoy.isoformat())
        con = sum(1 for _, _, t in patas if datos_patas.get(t, {}).get(hoy.isoformat(), {}).get("tea"))
        print(f"Patas de duales: {len(patas)} pedidas, {con} con tasa "
              f"(las dominadas vienen en null a propósito)")

    en_ccl = [dict(it, moneda="mep") for it in items if it.get("moneda") == "ccl"]
    homogeneos = {}
    if en_ccl:
        homogeneos = pedir_series(cli, en_ccl, ayer.isoformat(), hoy.isoformat())
        print(f"Segunda pasada en MEP para {len(en_ccl)} instrumentos que se valúan en CCL: "
              f"{len(homogeneos)} con dato")

    # RUEDAS DE REFERENCIA PARA LOS CIERRES. Se piden aparte y sólo el día que hacen falta: el
    # rango hoy-ayer no las cubre, y traer toda la semana o todo el mes multiplicaría los créditos
    # por cinco o por veinte para usar dos fechas.
    #
    # Se le piden a 1816 y no se sacan de historicos.xlsx aunque el Excel tenga la serie desde
    # diciembre: ahí sólo hay PRECIO, y el informe compara además tasas, paridades y durations. Con
    # el Excel el cierre semanal podría decir cuánto se movió el precio pero no cuánto la TIR, que
    # en pesos es justamente lo que se mira.
    refs = {}
    for tipo in ("semanal", "mensual"):
        if tipo not in tipos:
            continue
        f = ultima_rueda_de_periodo_anterior(hoy, fer, tipo)
        print(f"Cierre {tipo}: referencia {f}")
        # Se pide con un día de margen hacia atrás porque /series no devuelve nada cuando la fecha
        # inicial y la final coinciden; después se toma sólo la rueda que interesa.
        d2 = pedir_series(cli, items, (f - timedelta(days=4)).isoformat(), f.isoformat())
        refs[tipo] = {"fecha": f.isoformat(), "datos": d2}
        print(f"  {sum(1 for v in d2.values() if f.isoformat() in v)} tickers con dato en esa rueda")

    print(f"Leyes de ONs: {sum(1 for v in ons.values() if v['ley'] == 'local')} local, "
          f"{sum(1 for v in ons.values() if v['ley'] == 'ny')} NY")

    instrumentos, por_familia = [], defaultdict(list)
    sin_dato, sin_ley = [], []
    for it in items:
        fam = FAMILIAS.get(it["hoja"])
        if not fam or not it["t1816"]:
            continue
        if fam == "ONs":
            info = ons.get(it["eco"]) or {}
            fam = {"local": "ONs ley local", "ny": "ONs ley NY"}.get(info.get("ley"))
            if not fam:
                sin_ley.append(it["eco"])
                continue
        serie = datos.get(it["t1816"], {})
        h = serie.get(hoy.isoformat())
        a = serie.get(ayer.isoformat())
        if not h:
            sin_dato.append(it["eco"])
            continue
        reg = {
            "ticker": it["eco"],
            "familia": fam,
            "moneda": it["moneda"],
            "precio": h.get("precioDirty"),
            "tea": redondear(pct(h.get("tea"))),
            "durationMod": redondear(h.get("durationMod")),
            "paridad": redondear(pct(h.get("paridad"))),
            # Las dos varas: el % de precio, que es lo que ve el tenedor, y los puntos de tasa, que
            # es lo que se negocia. Para una LECAP la primera es casi siempre positiva por el mero
            # devengamiento, así que sin la segunda el informe diría que "todo subió" todos los días.
            "varPrecio": variacion(h.get("precioDirty"), (a or {}).get("precioDirty")),
            "varTasa": delta(pct(h.get("tea")), pct((a or {}).get("tea"))),
            "varParidad": delta(pct(h.get("paridad")), pct((a or {}).get("paridad"))),
            "conAyer": bool(a),
        }
        # Las patas, para los duales. `itm` marca la que hoy manda: es la única para la que 1816
        # calcula indicadores, así que tener tasa ES la señal.
        if fam == "Duales":
            pp = {}
            for tk, tipo, t1816 in patas:
                if tk != it["eco"]:
                    continue
                v = datos_patas.get(t1816, {}).get(hoy.isoformat())
                if not v:
                    continue
                pp[tipo] = {"tea": redondear(pct(v.get("tea"))),
                            "durationMod": redondear(v.get("durationMod")),
                            "itm": v.get("tea") is not None}
            if pp:
                reg["patas"] = pp

        # La punta homogénea, para poder compararlo con los que se valúan en MEP. Va como campo
        # aparte y no reemplaza al principal: la tabla por familia tiene que seguir mostrando lo
        # mismo que la pantalla.
        if it["moneda"] == "ccl":
            hm = homogeneos.get(it["t1816"], {}).get(hoy.isoformat())
            if hm:
                reg["enMep"] = {"precio": hm.get("precioDirty"),
                                "tea": redondear(pct(hm.get("tea"))),
                                "paridad": redondear(pct(hm.get("paridad")))}

        # Variación contra el cierre de la semana y del mes anteriores, con la misma vara que la
        # diaria: porcentaje de precio y puntos de tasa.
        for tipo, ref in refs.items():
            r0 = (datos_ref := ref["datos"].get(it["t1816"], {})).get(ref["fecha"])
            if r0:
                reg[f"varPrecio_{tipo}"] = variacion(h.get("precioDirty"), r0.get("precioDirty"))
                reg[f"varTasa_{tipo}"] = delta(pct(h.get("tea")), pct(r0.get("tea")))
        instrumentos.append(reg)
        por_familia[fam].append(reg)

    resumen = {}
    for fam, regs in por_familia.items():
        # OUTLIERS DE CONVENCIÓN. Dentro de una familia, 1816 no siempre devuelve la tasa en la
        # misma convención. Los duales CER/TAMAR son el caso claro: pagan el máximo entre CER más
        # margen y TAMAR, y el 2026-08-28 la API devolvía TXMD8 al 40,55% —TEA nominal— y TXMJ8 al
        # 5,71%, que sólo se entiende como TIR real. Promediarlas da un número que no significa
        # nada, y peor, un informe que afirma que los duales rinden 25%.
        #
        # No se corrigen ni se descartan: se marcan. Cuál de las dos convenciones es la correcta
        # depende del instrumento y no se puede resolver desde acá; lo que sí se puede es evitar que
        # el análisis los lea como comparables.
        teas = [r["tea"] for r in regs if r["tea"] is not None]
        raros = []
        if len(teas) >= 4:
            m = median(teas)
            # Factor 3 y no un múltiplo de desvío: lo que se busca no es un instrumento caro sino
            # una unidad distinta, y eso siempre aparece como un salto de orden de magnitud.
            raros = sorted(r["ticker"] for r in regs
                           if r["tea"] is not None and m > 0
                           and (r["tea"] > m * 3 or r["tea"] < m / 3))
        # La MEDIANA y no el promedio: en cada familia hay siempre algún instrumento ilíquido cuyo
        # precio quedó de hace tres ruedas y salta 4% cuando por fin opera. Con promedio, ese solo
        # dato define el signo de toda la familia.
        rp = resumir([r["varPrecio"] for r in regs])
        rt = resumir([r["varTasa"] for r in regs])
        con_var = [r for r in regs if r["varPrecio"] is not None]
        resumen[fam] = {
            "instrumentos": len(regs),
            "conVariacion": len(con_var),
            "metrica": METRICA.get(fam),
            "precio": rp,
            "tasa": rt,
            "teaMediana": resumir([r["tea"] for r in regs]),
            "mejor": max(con_var, key=lambda r: r["varPrecio"])["ticker"] if con_var else None,
            "peor": min(con_var, key=lambda r: r["varPrecio"])["ticker"] if con_var else None,
            # Tickers cuya tasa está en otra escala que el resto de su familia: hay que mirarlos
            # antes de compararlos con sus pares. Ver el comentario de arriba.
            "convencionDudosa": raros,
            # MONEDA EN LA QUE SE VALÚA LA FAMILIA, la misma que muestra el chip del monitor. No
            # siempre es una sola: las ONs de ley local son 54 en MEP y 7 en CCL, según en qué
            # moneda paga cada emisor. Por eso va el reparto completo y no un rótulo único, que
            # para esa familia sería mentira.
            "monedas": dict(Counter(r["moneda"] for r in regs).most_common()),
        }
        for tipo in refs:
            rp = resumir([r.get(f"varPrecio_{tipo}") for r in regs])
            rt = resumir([r.get(f"varTasa_{tipo}") for r in regs])
            if rp or rt:
                resumen[fam][tipo] = {"precio": rp, "tasa": rt}

    # BCRA, riesgo país y caución. Va después de los bonos y no antes porque si 1816 no responde el
    # informe no sale igual: sin precios no hay nada que contar, y estas series son el contexto.
    print("Pidiendo BCRA, riesgo país y caución...")
    macro = datos_macro(hoy, cliente_1816=cli,
                        referencias={t: r["fecha"] for t, r in refs.items()})
    if macro["fallos"]:
        print("  fallos macro:", "; ".join(macro["fallos"]))
    print(f"  {len(macro['series'])} series del BCRA · caución 1816: {macro['caucion']}")

    salida = {
        "fecha": hoy.isoformat(),
        "ruedaAnterior": ayer.isoformat(),
        "tipos": tipos,
        # EN HORA ARGENTINA. El runner de GitHub corre en UTC, así que datetime.now() daba las
        # 17:08 para una corrida de las 14:08 de Buenos Aires. El informe del 28/08/2026 salió
        # diciendo "tomados a las 17:08 con el mercado ya cerrado" cuando el mercado estaba abierto
        # y los precios eran intradía.
        "generado": datetime.now(ART).isoformat(timespec="seconds"),
        "feriadosLeidos": fer is not None,
        "universo": len(items),
        "sinDato": sin_dato,
        # ONs que no declaran ley en el Excel: quedan fuera del informe en vez de caer en una
        # familia arbitraria, y se listan para que se pueda completar la columna.
        "onsSinLey": sin_ley,
        "campos": CAMPOS,
        "resumen": resumen,
        "instrumentos": instrumentos,
        "macro": macro,
        # Fechas contra las que se midió cada cierre, para que el informe pueda nombrarlas en vez
        # de decir "la semana pasada".
        "referencias": {t: r["fecha"] for t, r in refs.items()},
    }

    DIR_INFORMES.mkdir(exist_ok=True)
    ruta = DIR_INFORMES / f"datos_{hoy.isoformat()}.json"
    ruta.write_text(json.dumps(salida, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"Escrito {ruta} · {len(instrumentos)} instrumentos · cierres: {', '.join(salida['tipos'])}")
    if sin_dato:
        print(f"Sin dato de hoy ({len(sin_dato)}): {', '.join(sin_dato[:15])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
