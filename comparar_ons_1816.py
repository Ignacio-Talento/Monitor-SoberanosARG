#!/usr/bin/env python3
"""
Compara, ON por ON, la TIR que calcula el monitor contra la que publica 1816.

POR QUÉ. De las 97 ONs del monitor, 88 tienen cronograma cargado y se calculan LOCALMENTE: el
monitor descuenta los flujos de la hoja Flujos contra el precio, y 1816 no interviene. Las otras 9
usan los indicadores de 1816 tal cual y coinciden por construcción. Entonces cualquier diferencia
sale de las 88, y puede venir de un cronograma mal cargado, de una amortización que no se reflejó,
o de una convención distinta.

Este script replica la cuenta del monitor (la rama onlocal/onny de calcMetricas en ons.html):
descuenta cada flujo posterior a la liquidación T+1 hábil, en base 365, contra el precio de la
última rueda de historicos.xlsx —que es el mismo precioDirty de 1816 que ve el frontend— y lo pone
al lado del `tea` que devuelve 1816 para esa fecha.

No modifica nada. Es un diagnóstico.

USO
    python comparar_ons_1816.py                 # todas las ONs con cronograma
    python comparar_ons_1816.py IRCFD MR44D     # sólo esas
    python comparar_ons_1816.py --umbral 1.0    # sólo las que difieran más de 1 pp
"""

import sys
from datetime import date, datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl", file=sys.stderr)
    sys.exit(1)

from actualizar_historicos import HISTORICOS_FILE, INSTRUMENTOS_FILE

try:
    from precios_1816 import Cliente1816, Error1816, MAX_TICKERS_SERIES
except ImportError:
    print("ERROR: no se pudo importar precios_1816", file=sys.stderr)
    sys.exit(1)

ESPERAS = [15, 45, 90]
UMBRAL_DEFECTO = 0.0        # en puntos porcentuales; 0 = mostrar todas


def a_fecha(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


def habil(d, feriados=()):
    return d.weekday() < 5 and d not in feriados


def sumar_habiles(d, n):
    while n > 0:
        d += timedelta(days=1)
        if habil(d):
            n -= 1
    return d


def pago_efectivo(d):
    """Mismo criterio que fechaPagoEfectiva() en el frontend: si cae inhábil, se corre adelante."""
    while not habil(d):
        d += timedelta(days=1)
    return d


def tir(flujos, precio):
    """Newton-Raphson sobre base 365, igual que calcTIR() en ons.html. -> TEA decimal."""
    r = 0.10
    for _ in range(300):
        vpn, dvpn = -precio, 0.0
        for t, monto in flujos:
            disc = (1 + r) ** t
            vpn += monto / disc
            dvpn -= t * monto / (disc * (1 + r))
        if abs(dvpn) < 1e-12:
            break
        rn = r - vpn / dvpn
        if abs(rn - r) < 1e-10:
            r = rn
            break
        r = max(rn, -0.999)
    return r


def leer_todo():
    wb = load_workbook(INSTRUMENTOS_FILE, data_only=True)
    ons = {}
    for r in wb["ONs"].iter_rows(values_only=True):
        if r and r[0] and str(r[0]).strip() != "Ticker":
            ons[str(r[0]).strip()] = {"nombre": str(r[1] or ""), "venc": a_fecha(r[2]),
                                      "ley": r[3]}
    filas = [r for r in wb["Flujos"].iter_rows(values_only=True) if r and r[0]]
    hdr = [str(c).strip() if c else "" for c in filas[0]]
    iT, iF, iTot = hdr.index("Ticker"), hdr.index("Fecha"), hdr.index("Total")
    iRes = hdr.index("Valor Residual") if "Valor Residual" in hdr else None
    flu = {}
    for r in filas[1:]:
        f = a_fecha(r[iF])
        if not f:
            continue
        flu.setdefault(str(r[iT]).strip(), []).append(
            (f, float(r[iTot] or 0), float(r[iRes] or 0) if iRes is not None else None))

    ws = load_workbook(HISTORICOS_FILE, read_only=True)["Historicos"]
    it = ws.iter_rows(values_only=True)
    cab = [str(c) for c in next(it)]
    ruedas = [r for r in it if r and r[0]]
    ult = ruedas[-1]
    fecha = str(ult[0])[:10]
    precio = {cab[i]: ult[i] for i in range(1, len(cab))
              if ult[i] not in (None, "") and isinstance(ult[i], (int, float))}
    return ons, flu, precio, fecha


def pedir(cli, tickers, fecha):
    """tea de 1816 para esa fecha. Reintenta ante 429: un lote caído no corta el resto."""
    import time
    y, m, d = map(int, fecha.split("-"))
    hasta = (date(y, m, d) + timedelta(days=1)).isoformat()   # /series necesita ventana, no un día
    out, ultimo = {}, None
    for i in range(0, len(tickers), MAX_TICKERS_SERIES):
        lote = tickers[i:i + MAX_TICKERS_SERIES]
        for espera in [0] + ESPERAS:
            if espera:
                print(f"    reintentando en {espera}s...", file=sys.stderr)
                time.sleep(espera)
            try:
                filas = cli.series(lote, ["tea"], moneda="mep",
                                   fecha_inicial=fecha, fecha_final=hasta)
                for r in filas:
                    if str(r.get("fecha", ""))[:10] == fecha and r.get("tea") is not None:
                        out[r["ticker"]] = r["tea"] * 100
                break
            except Error1816 as e:
                ultimo = e
                if "429" not in str(e) and "Demasiadas" not in str(e):
                    break
            except Exception as e:
                ultimo = e
        else:
            print(f"    lote {i//MAX_TICKERS_SERIES}: sin respuesta ({ultimo})", file=sys.stderr)
    return out


def main(argv):
    umbral = UMBRAL_DEFECTO
    if "--umbral" in argv:
        umbral = float(argv[argv.index("--umbral") + 1])
        argv = [a for i, a in enumerate(argv)
                if a != "--umbral" and (i == 0 or argv[i - 1] != "--umbral")]
    pedidos = {a.upper() for a in argv if not a.startswith("--")}

    ons, flu, precio, fecha = leer_todo()
    liq = sumar_habiles(date(*map(int, fecha.split("-"))), 1)
    print(f"rueda {fecha} · liquidación T+1 {liq}\n")

    objetivo = sorted(t for t in ons if t in flu and t in precio
                      and (not pedidos or t in pedidos))
    if not objetivo:
        print("Nada para comparar.")
        return 1

    cli = Cliente1816()
    ref = pedir(cli, [t[:-1] + "O" for t in objetivo], fecha)

    print(f"{'ticker':7s} {'precio':>8s} {'resid':>6s} {'monitor':>9s} {'1816':>9s} "
          f"{'dif pp':>8s}  nombre")
    print("-" * 78)
    filas = []
    for t in objetivo:
        p = precio[t]
        vivos = [(f, m, res) for f, m, res in flu[t] if f > liq]
        if not vivos:
            continue
        residual = max((res for f, m, res in flu[t] if f <= liq and res is not None),
                       default=None)
        fl = [((pago_efectivo(f) - liq).days / 365, m) for f, m, _ in vivos]
        mon = tir(fl, p) * 100
        r18 = ref.get(t[:-1] + "O")
        dif = (mon - r18) if r18 is not None else None
        if dif is None or abs(dif) >= umbral:
            filas.append((abs(dif) if dif is not None else -1, t, p, residual, mon, r18, dif))
    filas.sort(reverse=True)
    for _, t, p, res, mon, r18, dif in filas:
        rs = f"{res:6.1f}" if res is not None else "     —"
        r18s = f"{r18:8.2f}%" if r18 is not None else "       —"
        difs = f"{dif:+8.2f}" if dif is not None else "       —"
        print(f"{t:7s} {p:8.2f} {rs} {mon:8.2f}% {r18s} {difs}  {ons[t]['nombre'][:22]}")
    print(f"\n{len(filas)} filas · sin dato de 1816: {sum(1 for f in filas if f[5] is None)}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
