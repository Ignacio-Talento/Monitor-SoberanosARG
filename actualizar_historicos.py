"""
Script para actualizar historicos.xlsx con precios de cierre del día.
Se ejecuta automáticamente via GitHub Actions al cierre del mercado.
Los tickers se leen dinámicamente desde Instrumentos.xlsx.
"""

import requests
import openpyxl
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import os
import re
import time
from statistics import median

# Cliente de la API de 1816 (fuente primaria de precios). Import defensivo: si
# falta el archivo o la librería, el script sigue funcionando 100% con Eco Valores.
try:
    from precios_1816 import Cliente1816
except Exception as _e:
    Cliente1816 = None
    print(f"AVISO: cliente 1816 no disponible ({_e}); se usará solo Eco Valores.")

# ── CONFIGURACIÓN ─────────────────────────────────────────────
ECO_BASE        = "https://bonos.ecovalores.com.ar/eco/ticker.php"
HISTORICOS_FILE = "historicos.xlsx"
INSTRUMENTOS_FILE = "Instrumentos.xlsx"

# Grupos que usan sufijo D para buscar el precio en USD
GRUPOS_CON_D = {'USD Bonares', 'USD Globales'}

# ── MAPEO A LA API DE 1816 ────────────────────────────────────
# El valor que guarda Eco == 1816 'precioDirty' (dirty ya incorpora el residual
# de los amortizables). Moneda por hoja: 'ars' para instrumentos en pesos, 'mep'
# para los que Eco guarda en dólares (especie D). Verificado contra la API real.
CAMPO_1816 = "precioDirty"
MONEDA_1816 = {
    'LECAPS': 'ars', 'TASA FIJA': 'ars', 'CER': 'ars', 'TAMAR': 'ars',
    'USD Linked': 'ars', 'Duales': 'ars',
    'USD Bonares': 'mep',
    # Globales al CCL: es como los valúa la solapa "dólares" del monitor, y el histórico tiene que
    # estar en la MISMA moneda o las variaciones (DAY/WTD/MTD/YTD) comparan puntas distintas.
    'USD Globales': 'ccl',
    'USD Bopreales': 'mep', 'ON USD': 'mep',
    # Solapa ONs (ley local + NY): tickers en forma D, se piden a 1816 en mep con swap D->O.
    'ONs': 'mep',
    # Subsoberanos (provinciales USD): ticker idéntico en 1816. Al CCL, igual que los globales y
    # por el mismo motivo. Ojo: NUNCA a Eco —devuelve pesos o volumen y rompe los retornos—, para
    # eso están en HOJAS_SIN_ECO.
    'Subsoberanos': 'ccl',
}
# Hojas que NO son listas de instrumentos. 'Flujos' es la tabla de cronogramas y también tiene
# columna 'Ticker', así que se colaba entera: 17 tickers que no están en ninguna hoja de
# instrumentos (PUA36, BC37D, BDC33 y 14 más, varios ya vencidos) entraban al histórico. Y como
# 'Flujos' no está en MONEDA_1816, no mapeaban a 1816 y caían a Eco, que los devuelve en PESOS:
# la serie quedaba ~1.500x más grande y las variaciones contra el precio MEP daban -99%.
# La guarda de escala no lo detectaba porque compara contra la mediana del propio ticker, que
# también estaba en pesos: la serie era consistentemente errónea, no un salto puntual.
HOJAS_NO_INSTRUMENTOS = {'Flujos'}
# Hojas que NO deben caer a Eco: Eco no las publica en dólares y devuelve el precio en PESOS,
# que entra al histórico ~1.500x más grande y destroza los retornos (una sola celda mala hace
# que el WTD/MTD del bono muestre -99%). Si 1816 no tiene el precio, se prefiere dejar el hueco.
HOJAS_SIN_ECO = {'Subsoberanos', 'ONs', 'ON USD'}
# Guarda de escala: un precio que se aparta de la mediana reciente del propio ticker por más de
# este factor es casi seguro otra moneda u otra unidad; se descarta en vez de contaminar la serie.
FACTOR_ESCALA = 5

# Bopreales: el ticker de 1816 es irregular (no es un simple swap), mapa explícito.
MAPA_BOPREAL_1816 = {
    # Patrón: BP{XX}D -> BPO{XX}. Se deja explícito por si alguna serie no lo respeta.
    'BPA7D': 'BPOA7', 'BPB7D': 'BPOB7', 'BPC7D': 'BPOC7', 'BPD7D': 'BPOD7',
    'BPA8D': 'BPOA8', 'BPB8D': 'BPOB8',
}

def medianas_recientes(ws, header, ruedas=20):
    """Mediana de las últimas `ruedas` cotizaciones guardadas de cada ticker.
    Sirve de referencia de escala: si el precio nuevo se aparta muchísimo, es otra moneda.
    Devuelve {} si la hoja todavía no tiene historia suficiente (no aplica la guarda)."""
    medianas = {}
    for ticker, col in header.items():
        vals = []
        for row in range(ws.max_row, 1, -1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)) and v > 0:
                vals.append(v)
                if len(vals) >= ruedas:
                    break
        if len(vals) >= 3:          # con 1-2 datos la mediana no es confiable
            vals.sort()
            medianas[ticker] = vals[len(vals) // 2]
    return medianas


# Moneda de cada ON, leída de la hoja ONs: {ticker: 'ccl'|'mep'}. Se arma una vez y se consulta
# desde resolver_1816. Hace falta porque la moneda de una ON NO depende de la hoja sino de dónde
# paga: las de ley NY van al CCL y las locales según su columna Divisa (hoy 6 en CCL y 54 en MEP).
#
# Y hace falta consultarlo para las ONs de CUALQUIER hoja, no sólo de "ONs". El orden de hojas del
# Excel pone "ON USD" antes, y como leer_tickers deduplica por ticker, esas 4 se quedaban con la
# moneda de esa hoja: TLCPD y GN49D son ley NY y habrían seguido pidiéndose en MEP.
_MONEDA_ON = None
def moneda_on(ticker):
    """-> 'ccl' o 'mep' para una ON, o None si no está en la hoja ONs."""
    global _MONEDA_ON
    if _MONEDA_ON is None:
        _MONEDA_ON = {}
        try:
            wb = load_workbook(INSTRUMENTOS_FILE, data_only=True)
            if 'ONs' in wb.sheetnames:
                filas = [r for r in wb['ONs'].iter_rows(values_only=True) if r and r[0]]
                cab = next((i for i, r in enumerate(filas)
                            if str(r[0]).strip() == 'Ticker'), None)
                if cab is not None:
                    hdr = [str(c).strip() if c else '' for c in filas[cab]]
                    iL = hdr.index('Ley') if 'Ley' in hdr else None
                    iD = hdr.index('Divisa') if 'Divisa' in hdr else None
                    for r in filas[cab + 1:]:
                        t = str(r[0]).strip()
                        if not t or t == 'Ticker':
                            continue
                        ley = str(r[iL]).strip().lower() if iL is not None and iL < len(r) else ''
                        div = str(r[iD]).strip().upper() if iD is not None and iD < len(r) else ''
                        _MONEDA_ON[t] = 'ccl' if (ley == 'ny' or div == 'CCL') else 'mep'
        except Exception as e:
            print(f"AVISO: no se pudo leer la moneda de las ONs ({e}); se usa la de la hoja")
    return _MONEDA_ON.get(ticker)


def resolver_1816(sheet_name, eco_ticker, master_ticker):
    """Devuelve (ticker_1816, moneda) para consultar 1816, o (None, None) si no aplica.
    - Bonares/Globales: 1816 usa el ticker del master (sin la 'D' que agrega Eco).
    - ON USD: swap de la 'D' final por 'O' (RUCED -> RUCEO).
    - Bopreales: mapa explícito.
    - Resto (pesos): mismo ticker.
    Cualquier caso no resuelto -> (None, None) => fallback a Eco.
    """
    moneda = MONEDA_1816.get(sheet_name)
    if moneda is None:
        return None, None
    if sheet_name in GRUPOS_CON_D:
        return master_ticker, moneda
    if sheet_name == 'USD Bopreales':
        return MAPA_BOPREAL_1816.get(eco_ticker), moneda
    if sheet_name in ('ON USD', 'ONs'):
        t = (eco_ticker[:-1] + 'O') if eco_ticker.endswith('D') else None
        return t, (moneda_on(eco_ticker) or moneda)
    return eco_ticker, moneda  # pesos: idéntico

# ── LEER TICKERS DESDE INSTRUMENTOS.XLSX ─────────────────────
def leer_tickers():
    """Devuelve una lista de dicts: {'eco', 't1816', 'moneda'} por instrumento.
    'eco' es la columna de historicos (igual que antes); 't1816'/'moneda' se usan
    para pedir el precio a 1816 (None si el instrumento no mapea a 1816)."""
    if not os.path.exists(INSTRUMENTOS_FILE):
        print(f"ERROR: No se encontró {INSTRUMENTOS_FILE}")
        return []

    # read_only=False a propósito. En modo read_only openpyxl confía en el <dimension> que declara
    # la hoja, y si ese rango quedó viejo IGNORA las filas que estén más abajo: D15E7 estaba cargado
    # desde el 31/07 en la fila 10 de "USD Linked" con el dimension en A1:F9, así que el job dejó de
    # pedirle precio el 10/08 sin decir nada. El archivo pesa 200 KB; leerlo entero no se nota, y
    # perder un instrumento en silencio sí.
    wb = load_workbook(INSTRUMENTOS_FILE, data_only=True)
    items = []
    vistos = set()

    for sheet_name in wb.sheetnames:
        if sheet_name in HOJAS_NO_INSTRUMENTOS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Buscar fila de header (contiene 'Ticker')
        header_row = None
        for i, row in enumerate(rows):
            if row and str(row[0]).strip() == 'Ticker':
                header_row = i
                break

        if header_row is None:
            continue

        headers = [str(c).strip() if c else '' for c in rows[header_row]]
        ticker_col = headers.index('Ticker') if 'Ticker' in headers else 0

        for row in rows[header_row + 1:]:
            if not row or not row[ticker_col]:
                continue
            ticker = str(row[ticker_col]).strip()
            if not ticker or ticker == 'Ticker' or ticker == 'None':
                continue

            # Agregar sufijo D para Bonares y Globales
            eco_ticker = ticker + 'D' if sheet_name in GRUPOS_CON_D else ticker

            if eco_ticker in vistos:
                continue
            vistos.add(eco_ticker)

            t1816, moneda = resolver_1816(sheet_name, eco_ticker, ticker)
            items.append({'eco': eco_ticker, 't1816': t1816, 'moneda': moneda,
                          'hoja': sheet_name})

    print(f"Tickers leídos desde {INSTRUMENTOS_FILE}: {len(items)}")
    return items

# ── CLIENTE 1816 COMPARTIDO ───────────────────────────────────
# Una sola instancia para todo el proceso: el cliente lleva internamente el control de
# 1 request/seg que exige 1816, así que crear uno por función hace que cada uno espacie
# por su cuenta, se pisen entre sí y la API devuelva HTTP 429.
_CLI_1816 = None
def cliente_1816():
    """Devuelve el cliente compartido, o None si no hay key/cliente disponible."""
    global _CLI_1816
    if Cliente1816 is None:
        return None
    if not (os.environ.get("API_1816_KEY") or os.path.exists(".1816_key")):
        return None
    if _CLI_1816 is None:
        _CLI_1816 = Cliente1816()
    return _CLI_1816

# ── FETCH PRECIOS DESDE 1816 (fuente primaria) ────────────────
def fetch_precios_1816(items, fecha=None):
    """Devuelve {eco_ticker: precio} solo para los que 1816 respondió con dato.
    `fecha` (AAAA-MM-DD) es opcional: por defecto usa el día de hoy (producción);
    se puede fijar para pruebas o backfills puntuales.
    Ante cualquier problema (sin key, sin cliente, error de red/API) devuelve {}
    y el flujo cae a Eco Valores para todo. Nunca rompe la corrida."""
    cli = cliente_1816()
    if cli is None:
        print("AVISO: no hay API_1816_KEY o cliente 1816; se usará solo Eco Valores.")
        return {}

    # Agrupar por moneda: {moneda: [(eco, t1816), ...]}
    por_moneda = {}
    for it in items:
        if it['t1816'] and it['moneda']:
            por_moneda.setdefault(it['moneda'], []).append((it['eco'], it['t1816']))

    if not por_moneda:
        return {}

    resultado = {}
    fallaron = []
    for moneda, pares in por_moneda.items():
        tickers = [t for _, t in pares]
        # Esperas largas a propósito. El limitador de 1816 es GLOBAL por API key y el job del repo
        # del colega corría con el MISMO cron que éste, así que los dos empezaban juntos y el que
        # llegaba segundo se comía el 429. Como ese job tarda uno o dos minutos en soltar la cuota,
        # reintentar a los 3 y 6 segundos no servía de nada: había que esperar en esa escala.
        # (El cron de este repo se corrió 15 minutos para no depender sólo de esto.)
        filas = None
        for espera in (0, 15, 45, 90):
            if espera:
                print(f"  {moneda}: reintentando en {espera}s...")
                time.sleep(espera)
            try:
                filas = cli.precios(tickers, [CAMPO_1816], moneda=moneda, fecha_operacion=fecha)
                break
            except Exception as e:
                ultimo = e
        if filas is None:
            # Este grupo se pierde, pero los que YA se trajeron se conservan: antes un fallo en el
            # último grupo descartaba todo lo anterior y mandaba los 171 tickers a Eco, que no
            # cubre subsoberanos ni ONs. Así se perdían ~95 precios por un rate limit transitorio
            # (pasó el 12 y el 14 de agosto de 2026: 71 y 70 tickers guardados contra los ~165
            # habituales). Ahora Eco cubre sólo los huecos reales.
            print(f"AVISO: 1816 falló en {moneda} ({ultimo}); esos tickers van a Eco.")
            fallaron.append(moneda)
            continue
        valor_por_t = {f['ticker']: f.get(CAMPO_1816) for f in filas}
        for eco, t in pares:
            v = valor_por_t.get(t)
            if isinstance(v, (int, float)):
                resultado[eco] = v

    # Los que se piden en CCL y no operaron en esa punta: se traen en MEP y se convierten.
    faltanCCL = [(it['eco'], it['t1816']) for it in items
                 if it['moneda'] == 'ccl' and it['t1816'] and it['eco'] not in resultado]
    if faltanCCL and 'ccl' not in fallaron:
        factor = canje_ccl_mep(fecha or hoy_art().strftime('%Y-%m-%d'))
        if factor:
            try:
                filas = cli.precios([t for _, t in faltanCCL], [CAMPO_1816],
                                    moneda='mep', fecha_operacion=fecha)
                v_por_t = {f['ticker']: f.get(CAMPO_1816) for f in filas}
                n = 0
                for eco, t in faltanCCL:
                    v = v_por_t.get(t)
                    if isinstance(v, (int, float)) and v > 0:
                        resultado[eco] = round(v * factor, 4)
                        n += 1
                if n:
                    print(f"  {n} sin cotización en CCL: convertidos desde MEP "
                          f"(canje {factor:.4f})")
            except Exception as e:
                print(f"AVISO: no se pudieron traer los faltantes en MEP ({e})")

    total = sum(len(v) for v in por_moneda.values())
    extra = f" (falló: {', '.join(fallaron)})" if fallaron else ""
    print(f"1816: {len(resultado)} precios obtenidos de {total} consultables{extra}.")
    return resultado


# ── CANJE CCL/MEP ─────────────────────────────────────────────
# El CCL es una punta mucho más fina que el MEP: hay papeles —provinciales y ONs chicas— con precio
# MEP todos los días y CCL uno de cada tres, o ninguno. Como el monitor los valúa en CCL, pedirlos
# sólo en esa punta dejaba huecos: la rueda del 24-08 quedó con 17 sin precio contra los 3
# habituales, y 15 eran de este grupo.
#
# La salida es convertir: se les pide el precio en MEP y se multiplica por el canje del día. No es
# inventar un precio, es reexpresar el mismo valor en la otra moneda.
#
# El canje sale del índice dólar de BYMA, que publica MEP y CCL de la misma canasta. Se eligió esa
# fuente y no derivarlo de los bonos porque es gratis —no gasta créditos de 1816— y una sola
# llamada resuelve el día. Contrastado el 25-08-2026 contra el factor implícito en los ~44 bonos
# que cotizan en las dos puntas: coinciden dentro del 0,1%.
BYMA_AJAX = ("https://data-widgets.byma.com.ar/wp-admin/admin-ajax.php?action=get_indice_dolar")
BYMA_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0",
    "X-Requested-With": "XMLHttpRequest", "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors", "Sec-Fetch-Site": "same-origin",
    # Sin los Sec-Fetch-* y el Referer, el host contesta 401 aunque la URL sea correcta.
    "Referer": "https://data-widgets.byma.com.ar/indice-dolar-historico-widget/",
}


# Testigos para derivar el canje de 1816: bonos que cotizan TODOS los días en las dos puntas.
# Soberanos hard-dollar líquidos y de vencimiento repartido, para que la baja de uno no deje al
# cálculo sin base. Con que respondan 3 alcanza.
TESTIGOS_CANJE = ['AL30', 'GD30', 'AL29', 'GD29', 'AL35', 'GD35', 'AE38', 'GD38', 'AL41', 'GD41']
# Fuera de esta banda el cociente no es un canje sino otra cosa (un precio en pesos, un dato
# viejo). Se descarta el par en vez de arrastrarlo a la mediana.
CANJE_MIN, CANJE_MAX = 0.85, 1.30


def canje_desde_1816(fecha):
    """Canje CCL/MEP derivado de los testigos, o None.

    POR QUÉ ES LA FUENTE PRIMARIA Y NO BYMA. El endpoint de BYMA da SSL
    CERTIFICATE_VERIFY_FAILED desde Python —manda la cadena incompleta y no incluye el
    intermedio—. Los navegadores y los Workers de Cloudflare lo resuelven solos buscando el
    certificado que falta (AIA fetching); requests no hace eso, así que falla igual en Windows
    que en el runner de Ubuntu. El 25/08/2026 eso dejó 18 instrumentos en CCL sin precio en el
    histórico, en silencio, porque el aviso era sólo un print.

    Derivarlo de 1816 saca una dependencia externa del camino crítico: es la misma API que ya
    trae todo lo demás, con la misma key y el mismo rate limit. Y el número es el mismo — el
    factor de los ~44 instrumentos que cotizan en ambas puntas coincide con el índice BYMA
    dentro del 0,1%.

    Cuesta 20 créditos: 10 testigos x 1 campo x 2 monedas.
    """
    cli = cliente_1816()
    if cli is None:
        return None
    # Con reintentos, igual que fetch_precios_1816: el limitador de 1816 es global por API key, y
    # acá un 429 no degrada el resultado sino que lo borra —sin factor no se completa NINGÚN CCL—.
    def punta(moneda):
        ultimo = None
        for espera in (0, 15, 45):
            if espera:
                print(f"  canje/{moneda}: reintentando en {espera}s...")
                time.sleep(espera)
            try:
                filas = cli.precios(TESTIGOS_CANJE, [CAMPO_1816], moneda=moneda,
                                    fecha_operacion=fecha)
                return {f['ticker']: f.get(CAMPO_1816) for f in filas}
            except Exception as e:
                ultimo = e
        print(f"AVISO: no se pudo derivar el canje de 1816 en {moneda} ({ultimo})")
        return None

    ccl, mep = punta('ccl'), punta('mep')
    if ccl is None or mep is None:
        return None

    razones = []
    for t in TESTIGOS_CANJE:
        a, b = ccl.get(t), mep.get(t)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)) and b > 0:
            r = a / b
            if CANJE_MIN <= r <= CANJE_MAX:
                razones.append(r)
    if len(razones) < 3:
        print(f"AVISO: sólo {len(razones)} testigo(s) con las dos puntas; se prueba con BYMA")
        return None
    razones.sort()
    f = median(razones)
    print(f"  canje CCL/MEP {f:.4f} (1816, {len(razones)} testigos, "
          f"{razones[0]:.4f}-{razones[-1]:.4f})")
    return f


def canje_ccl_mep(fecha):
    """Factor para pasar un precio de MEP a CCL en esa rueda, o None si no se pudo.

    Primero 1816, que es la fuente que ya usa todo el script; BYMA queda de respaldo por si
    algún día los testigos no operan en CCL (ver canje_desde_1816 para el porqué del orden)."""
    f = canje_desde_1816(fecha)
    if f:
        return f
    try:
        r = requests.get(BYMA_AJAX, headers=BYMA_HEADERS, timeout=30)
        if not r.ok:
            print(f"AVISO: índice BYMA HTTP {r.status_code}; no se completa por canje")
            return None
        filas = {str(x.get("date", ""))[:10]: x for x in (r.json().get("result") or [])}
        x = filas.get(fecha)
        if not x:
            # Si la rueda todavía no está publicada se usa la última anterior: el canje se mueve
            # décimas por día, así que el error es de otro orden que dejar el precio afuera.
            previas = sorted(f for f in filas if f <= fecha)
            x = filas.get(previas[-1]) if previas else None
            if x:
                print(f"AVISO: BYMA todavía no publicó {fecha}; se usa el canje del {previas[-1]}")
        mep, ccl = (x or {}).get("bymaClosingPrice"), (x or {}).get("cclClosingPrice")
        if not isinstance(mep, (int, float)) or not isinstance(ccl, (int, float)) or not ccl:
            return None
        return mep / ccl
    except Exception as e:
        print(f"AVISO: no se pudo leer el índice BYMA ({e}); no se completa por canje")
        return None


# ── FECHA DE LA RUEDA A REGISTRAR ─────────────────────────────
# Referencias para detectar si una rueda tuvo mercado. Deben ser instrumentos MUY líquidos y
# de vencimiento lejano: si la referencia vence, deja de tener datos y la resolución se clava
# para siempre (le pasó a S31L6, que venció el 31-07-2026 y congeló el histórico 3 ruedas).
# Se usan varias por si alguna no opera ese día; alcanza con que una tenga precio.
REFERENCIAS_RUEDA = [('AL30', 'mep'), ('GD30', 'mep'), ('TZXO6', 'ars')]


def referencias_rueda(items):
    """Referencias a usar para detectar la última rueda con mercado.

    Prioriza las de REFERENCIAS_RUEDA que estén en el universo; si ninguna está (el Excel
    cambió mucho), cae a los primeros instrumentos mapeados, como antes."""
    disponibles = {it['t1816'] for it in items if it['t1816']}
    refs = [{'t1816': t, 'moneda': m} for t, m in REFERENCIAS_RUEDA if t in disponibles]
    if refs:
        return refs
    return [it for it in items if it['t1816'] and it['moneda']][:3]


def hoy_art():
    """Fecha de hoy en Argentina (UTC-3).

    El runner de GitHub Actions corre en UTC: una corrida de las 21:52 ART cae en el día
    siguiente según date.today(), y ahí arrancaba a buscar la rueda por un día que todavía
    no existía. Acá el calendario que importa es el argentino.
    """
    return (datetime.utcnow() - timedelta(hours=3)).date()


def resolver_fecha_1816(items, max_dias=7):
    """Última rueda con datos en 1816, buscando hacia atrás desde hoy.

    1816 no tiene datos los fines de semana, los feriados, ni antes del cierre:
    pedirle "hoy" a las 10 AM devuelve todo null. Sin esto la corrida degradaba a
    Eco en silencio y escribía una fila con los precios del día anterior, que además
    bloqueaba la corrida real de ese día (el chequeo de "ya existe fila" la saltea).

    Se prueba con UN ticker de referencia (barato) y se saltean sábados y domingos
    por fecha, sin gastar consultas. Devuelve 'AAAA-MM-DD', o None si no hay key /
    cliente / 1816 no responde (en ese caso el flujo sigue como antes: hoy + Eco).
    """
    cli = cliente_1816()
    if cli is None:
        return None
    refs = referencias_rueda(items)
    if not refs:
        return None
    try:
        for i in range(max_dias + 1):
            d = hoy_art() - timedelta(days=i)
            if d.weekday() >= 5:          # sábado/domingo: ni consultamos
                continue
            f = d.strftime("%Y-%m-%d")
            # Alcanza con que UNA de las referencias haya operado ese día.
            for ref in refs:
                # 1816 admite 1 req/seg: un 429 transitorio no debe degradar todo el día a Eco.
                for intento in range(3):
                    try:
                        filas = cli.precios([ref['t1816']], [CAMPO_1816],
                                            moneda=ref['moneda'], fecha_operacion=f)
                        break
                    except Exception:
                        if intento == 2:
                            raise
                        time.sleep(2 * (intento + 1))
                if filas and isinstance(filas[0].get(CAMPO_1816), (int, float)):
                    return f
    except Exception as e:
        print(f"AVISO: no se pudo resolver la fecha en 1816 ({e}).")
    return None

# ── FETCH PRECIO ──────────────────────────────────────────────
def fetch_precio(ticker):
    try:
        url = f"{ECO_BASE}?t={ticker}"
        resp = requests.get(url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html",
            "Referer": "https://bonos.ecovalores.com.ar"
        })
        html = resp.text
        match = re.search(r'<td class="precioticker">\s*([\d.,]+)\s*</td>', html)
        if match:
            price_str = match.group(1).replace(".", "").replace(",", ".")
            return float(price_str)
    except Exception as e:
        print(f"  Error fetching {ticker}: {e}")
    return None

# ── ACTUALIZAR EXCEL ──────────────────────────────────────────
def actualizar_historicos():
    # Leer tickers dinámicamente
    items = leer_tickers()
    if not items:
        print("ERROR: No se pudieron leer los tickers.")
        return
    tickers = [it['eco'] for it in items]

    # La fila se rotula con la última rueda que 1816 tenga cargada, que no siempre es
    # hoy (fin de semana, feriado, o corrida antes del cierre). Así una corrida
    # prematura no inventa una fila con precios viejos ni bloquea la corrida real:
    # si esa rueda ya está registrada, se sale por el chequeo de más abajo.
    fecha_1816 = resolver_fecha_1816(items)
    if fecha_1816:
        fecha_str = fecha_1816
        if fecha_str != hoy_art().strftime("%Y-%m-%d"):
            print(f"1816 todavía no tiene datos de hoy; última rueda disponible: {fecha_str}")
    elif cliente_1816() is not None:
        # Hay cliente de 1816 pero no se pudo resolver la rueda (típicamente un 429). Antes se
        # rotulaba la fila con date.today() y se pedían los precios sin fechaOperacion, lo que
        # escribe mal dos veces: (a) el runner de GitHub corre en UTC, así que después de las
        # 21 ART "hoy" ya es el día siguiente, y (b) 1816 sin fecha devuelve la última rueda
        # disponible, que es la anterior. El 2026-08-06 pasó exactamente eso: la corrida guardó
        # los cierres del 06 rotulados como 07, las variaciones de ese día daban cero y el
        # chequeo de "ya existe fila" bloqueaba la corrida real del 07.
        # Una fila faltante se recupera con el backfill; una mal rotulada corrompe en silencio.
        print("ERROR: 1816 está disponible pero no se pudo resolver la rueda (¿429?). No se "
              "escribe nada para no rotular mal la fila; correr el backfill para esa fecha.")
        return
    else:
        # Sin cliente de 1816 (falta la key o la librería): modo Eco puro, que es el
        # comportamiento histórico. Eco cotiza en vivo, así que la fecha de hoy sí corresponde
        # a los precios que devuelve. Se usa el calendario argentino, no el del runner.
        fecha_str = hoy_art().strftime("%Y-%m-%d")
        print("AVISO: 1816 no disponible; se usa la fecha de hoy (ARG) y se completa con Eco.")
    print(f"Actualizando historicos para {fecha_str}...")

    # Cargar o crear el Excel
    if os.path.exists(HISTORICOS_FILE):
        wb = load_workbook(HISTORICOS_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historicos"
        ws.cell(row=1, column=1, value="Fecha")
        print("Archivo historicos.xlsx creado desde cero.")

    # Verificar si ya existe la fila de hoy
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0])[:10] == fecha_str:
            print(f"Ya existe fila para {fecha_str}, saliendo.")
            return

    # Próxima fila vacía
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=fecha_str)

    # Asegurar que todos los tickers estén en el header
    header = {ws.cell(row=1, column=c).value: c for c in range(2, ws.max_column + 1)}
    for ticker in tickers:
        if ticker not in header:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=ticker)
            header[ticker] = new_col
            print(f"  Nuevo ticker agregado al header: {ticker}")

    # Precios primero desde 1816 (fuente primaria); lo que falte, desde Eco.
    precios_api = fetch_precios_1816(items, fecha=fecha_1816)

    # Mediana reciente por ticker (últimas ruedas ya guardadas), para la guarda de escala.
    medianas = medianas_recientes(ws, header)

    n1816 = 0
    neco = 0
    err = 0
    descartados = 0
    for it in items:
        ticker = it['eco']
        precio = precios_api.get(ticker)
        if precio is not None:
            fuente = "1816"
        elif it.get('hoja') in HOJAS_SIN_ECO:
            # Eco devolvería el precio en pesos para estos: mejor dejar el hueco.
            precio, fuente = None, "sin-eco"
        else:
            # Fallback: scraping de Eco Valores (comportamiento original).
            print(f"  Fetching {ticker} (Eco)...", end=" ")
            precio = fetch_precio(ticker)
            print(f"${precio}" if precio else "sin precio")
            time.sleep(0.4)  # throttle solo cuando efectivamente pegamos a Eco
            fuente = "eco"

        # Guarda de escala: descartar valores que no pueden ser el mismo instrumento.
        med = medianas.get(ticker)
        if precio and med and (precio > med * FACTOR_ESCALA or precio < med / FACTOR_ESCALA):
            print(f"  DESCARTADO {ticker}: {precio} vs mediana {med:.2f} ({fuente}) — ¿otra moneda?")
            precio = None
            descartados += 1

        if precio:
            ws.cell(row=next_row, column=header[ticker], value=precio)
            if fuente == "1816":
                n1816 += 1
            else:
                neco += 1
        else:
            err += 1

    wb.save(HISTORICOS_FILE)
    print(f"\nListo: {n1816 + neco} precios guardados "
          f"(1816: {n1816}, Eco: {neco}), {err} sin precio.")

if __name__ == "__main__":
    actualizar_historicos()
