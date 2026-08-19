#!/usr/bin/env python3
"""
Rellena la historia de los instrumentos recién dados de alta en historicos.xlsx.

POR QUÉ EXISTE. El job diario sólo escribe la rueda del día: un instrumento que se agrega hoy
arranca con una sola cotización y las columnas DAY/WTD/MTD/YTD del monitor le muestran "—" hasta
que pasen semanas. El precio está, pero no hay contra qué compararlo. Este script le pide a 1816
la serie completa y llena las ruedas que ya existen en el archivo.

Es el eje opuesto a completar_ruedas.py: aquel rellena una RUEDA a la que le faltan tickers (se
rompió la corrida de un día), éste rellena un TICKER al que le faltan ruedas (se dio de alta
tarde). Comparten el archivo y el criterio de no pisar nunca un valor que ya está.

QUÉ CUESTA. El costo de /series es tickers x campos x días CON DATO, así que un bono emitido en
julio cuesta las pocas ruedas que vivió y no el rango entero que se pide. Se pide una sola ventana
por moneda, no una por ticker.

USO
    python rellenar_historia.py                  # detecta los que tienen poca historia
    python rellenar_historia.py PNXCD VSCYD      # sólo esos (en la forma del monitor)
    python rellenar_historia.py --dry-run        # qué haría, sin gastar
"""

import sys
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl", file=sys.stderr)
    sys.exit(1)

from actualizar_historicos import CAMPO_1816, HISTORICOS_FILE, leer_tickers

try:
    from precios_1816 import Cliente1816, Error1816, MAX_TICKERS_SERIES
except ImportError:
    print("ERROR: no se pudo importar precios_1816", file=sys.stderr)
    sys.exit(1)

# Con menos ruedas que esto se considera que al instrumento le falta historia. Diez es bastante
# menos que un mes de mercado: un bono con historia real supera ese número aunque sea ilíquido, y
# uno recién cargado no llega ni cerca.
UMBRAL_HISTORIA = 10

# Reintentos ante 429. El limitador de 1816 es global por API key, así que si otro proceso está
# consumiendo hay que esperar de verdad.
ESPERAS = [15, 45, 90]


def leer(ws):
    """-> (cabecera, [fechas en orden], {fecha: {ticker: valor}}, {fecha: nro de fila})"""
    cab = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    fechas, datos, fila_de = [], {}, {}
    for r in range(2, ws.max_row + 1):
        f = str(ws.cell(row=r, column=1).value or "")[:10]
        if len(f) != 10 or f[4] != "-":
            continue
        fechas.append(f)
        fila_de[f] = r
        datos[f] = {cab[c - 1]: ws.cell(row=r, column=c).value
                    for c in range(2, len(cab) + 1)
                    if ws.cell(row=r, column=c).value not in (None, "")}
    fechas.sort()
    return cab, fechas, datos, fila_de


def pedir(cli, tickers, moneda, desde, hasta):
    """-> [filas] de 1816, o None si no contestó. Un lote que falla no corta el resto."""
    import time
    ultimo = None
    for espera in [0] + ESPERAS:
        if espera:
            print(f"      reintentando en {espera}s...")
            time.sleep(espera)
        try:
            return cli.series(tickers, [CAMPO_1816], moneda=moneda,
                              fecha_inicial=desde, fecha_final=hasta)
        except Error1816 as e:
            ultimo = e
            if "429" not in str(e) and "Demasiadas" not in str(e):
                break            # no es rate limit: reintentar no ayuda
        except Exception as e:
            ultimo = e
    print(f"      1816 falló ({ultimo}), se deja como estaba", file=sys.stderr)
    return None


def main(argv):
    dry = "--dry-run" in argv
    pedidos = {a.upper() for a in argv if not a.startswith("--")}

    wb = load_workbook(HISTORICOS_FILE)
    ws = wb["Historicos"]
    cab, fechas, datos, fila_de = leer(ws)
    if not fechas:
        print("El archivo no tiene ruedas.")
        return 1
    col = {t: i + 1 for i, t in enumerate(cab) if t}

    # Sólo lo que mapea a 1816 y está en la cabecera del histórico.
    items = [it for it in leer_tickers()
             if it["t1816"] and it["moneda"] and it["eco"] in col]
    if pedidos:
        objetivo = [it for it in items if it["eco"] in pedidos]
        for t in sorted(pedidos - {it["eco"] for it in objetivo}):
            print(f"AVISO: {t} no está en el histórico o no mapea a 1816, se saltea",
                  file=sys.stderr)
    else:
        cuenta = {it["eco"]: sum(1 for f in fechas if it["eco"] in datos[f]) for it in items}
        objetivo = [it for it in items if cuenta[it["eco"]] < UMBRAL_HISTORIA]

    if not objetivo:
        print("No hay instrumentos con historia faltante.")
        return 0

    print(f"{len(objetivo)} instrumentos a rellenar · ruedas del archivo: "
          f"{fechas[0]} a {fechas[-1]}")
    for it in sorted(objetivo, key=lambda x: x["eco"]):
        n = sum(1 for f in fechas if it["eco"] in datos[f])
        print(f"  {it['eco']:7s} -> 1816 {it['t1816']:7s} ({it['moneda']}) · {n} ruedas hoy")
    if dry:
        print("(el costo real son los días CON dato de cada uno, no el rango entero)")
        return 0

    cli = Cliente1816()
    inv = {it["t1816"]: it["eco"] for it in objetivo}
    porm = {}
    for it in objetivo:
        porm.setdefault(it["moneda"], []).append(it["t1816"])

    escritos = 0
    for moneda, lote in porm.items():
        for i in range(0, len(lote), MAX_TICKERS_SERIES):
            trozo = lote[i:i + MAX_TICKERS_SERIES]
            print(f"\n{moneda}: {len(trozo)} tickers, {fechas[0]} a {fechas[-1]}")
            filas = pedir(cli, trozo, moneda, fechas[0], fechas[-1])
            if filas is None:
                continue
            puestos = 0
            for r in filas:
                v = r.get(CAMPO_1816)
                eco = inv.get(r.get("ticker"))
                f = str(r.get("fecha", ""))[:10]
                if v is None or not eco or f not in fila_de:
                    continue           # sin dato, ticker ajeno, o rueda que el archivo no tiene
                if not isinstance(v, (int, float)) or v <= 0:
                    continue
                celda = ws.cell(row=fila_de[f], column=col[eco])
                if celda.value not in (None, ""):
                    continue           # no se pisa nada que ya esté
                celda.value = v
                puestos += 1
            print(f"      {len(filas)} filas de 1816, {puestos} celdas nuevas")
            escritos += puestos

    if escritos:
        wb.save(HISTORICOS_FILE)
        print(f"\n{HISTORICOS_FILE} actualizado: {escritos} precios")
    else:
        print("\nSin cambios.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
