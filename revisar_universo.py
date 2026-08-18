#!/usr/bin/env python3
"""
Compara el universo que publica 1816 contra los instrumentos que sigue el monitor.

Contesta las dos preguntas del repaso semanal: qué hay que dar de baja (venció) y qué habría que
dar de alta (1816 lo lista y el monitor no lo tiene). No modifica nada: el alta de verdad la
completa una persona, porque 1816 no trae cupones, margen ni lag y esos datos van a mano en
Instrumentos.xlsx.

Las curvas son las mismas que usa functions/api/instrumentos.js. Si allá se agrega una, va también
acá; si se separan, el detector del frontend y este script dejan de coincidir.

Las bajas se calculan sin gastar un crédito: el vencimiento ya está en Instrumentos.xlsx.

USO
    python revisar_universo.py            # repaso completo
    python revisar_universo.py --bajas    # sólo vencidos, sin tocar 1816
"""

import sys
from datetime import date, datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl", file=sys.stderr)
    sys.exit(1)

from actualizar_historicos import HOJAS_NO_INSTRUMENTOS, INSTRUMENTOS_FILE, leer_tickers

# curvaId de 1816 -> grupo del monitor. Espejo de CURVAS en functions/api/instrumentos.js.
CURVAS = {
    9: "lecap", 10: "tasafija", 7: "cer", 28: "tamar", 14: "dual",
    12: "usdlinked", 17: "usdlinked", 8: "usdbonares", 11: "usdglobales", 24: "usdbopreal",
    18: "subsoberano", 20: "subsoberano", 21: "subsoberano", 19: "subsoberano",
    15: "subsoberano", 27: "subsoberano", 29: "subsoberano",
    16: "onusd", 3: "oncorp", 25: "oncorp", 5: "oncorp", 4: "oncorp",
    26: "oncorp", 23: "oncorp", 30: "oncorp",
}

# Curvas que el monitor sigue de forma EXHAUSTIVA: si 1816 lista uno y no está, es un alta que
# falta. Las demás (provinciales y corporativos) son un subconjunto curado por liquidez, así que
# ahí un faltante no es un error: se listan aparte y con volumen, para decidir.
EXHAUSTIVAS = {"lecap", "tasafija", "cer", "tamar", "dual", "usdlinked",
               "usdbonares", "usdglobales", "usdbopreal"}

# Piso de volumen diario para que un corporativo o provincial valga la pena seguir. Mismo criterio
# que el auto-add del frontend.
PISO_VOLUMEN = 100_000


def hoy_art():
    return (datetime.utcnow() - timedelta(hours=3)).date()


def a_fecha(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and len(v) >= 10:
        try:
            return date(*map(int, v[:10].split("-")))
        except ValueError:
            return None
    return None


def leer_monitor():
    """-> {ticker: (hoja, vencimiento)} de todo lo que sigue el monitor.

    Las hojas de flujos no tienen columna de vencimiento: ahí el vencimiento es la última fecha de
    cashflow del ticker.
    """
    # Sin read_only: ver el comentario en leer_tickers(). Un <dimension> viejo esconde filas, y
    # este script justamente existe para que no se escape ningún instrumento.
    wb = load_workbook(INSTRUMENTOS_FILE, data_only=True)
    universo = {}
    for hoja in wb.sheetnames:
        if hoja in HOJAS_NO_INSTRUMENTOS:
            continue
        filas = list(wb[hoja].iter_rows(values_only=True))
        cab = next((i for i, f in enumerate(filas)
                    if f and str(f[0]).strip() == "Ticker"), None)
        if cab is None:
            continue
        hdr = [str(c).strip() if c else "" for c in filas[cab]]
        ivenc = next((i for i, h in enumerate(hdr) if h.lower().startswith("fecha venc")), None)
        flujos = ivenc is None
        if flujos:
            if "Fecha" not in hdr:
                continue
            ivenc = hdr.index("Fecha")
        for f in filas[cab + 1:]:
            if not f or not f[0]:
                continue
            t = str(f[0]).strip()
            if t in ("Ticker", "None"):
                continue
            venc = a_fecha(f[ivenc]) if ivenc < len(f) else None
            if flujos:                       # el vencimiento es el último cashflow
                anterior = universo.get(t, (hoja, None))[1]
                if venc and anterior and anterior > venc:
                    venc = anterior
            universo[t] = (hoja, venc)
    return universo


def volumenes(cli, tickers):
    """-> {ticker: volumen diario}. Un lote que falla no tira abajo el resto del repaso."""
    from precios_1816 import Error1816
    out = {}
    for i in range(0, len(tickers), 50):
        try:
            filas = cli.precios(tickers[i:i + 50], ["volumenMontoDiario"], moneda="mep")
        except Error1816 as e:
            print(f"AVISO: no se pudo medir volumen de un lote: {e}", file=sys.stderr)
            continue
        for f in filas:
            v = f.get("volumenMontoDiario")
            if isinstance(v, (int, float)):
                out[f["ticker"]] = v
    return out


def main(argv):
    hoy = hoy_art()
    monitor = leer_monitor()
    print(f"{len(monitor)} instrumentos en {INSTRUMENTOS_FILE} - hoy {hoy}\n")

    # --- BAJAS: gratis, sale del propio archivo ---
    vencidos = sorted((v, t, h) for t, (h, v) in monitor.items() if v and v <= hoy)
    print("=== BAJAS: ya vencieron y siguen en el archivo ===")
    for v, t, h in vencidos:
        print(f"  {v}  {h:15s} {t}")
    if not vencidos:
        print("  (ninguna)")

    proximos = sorted((v, t, h) for t, (h, v) in monitor.items()
                      if v and hoy < v <= hoy + timedelta(days=30))
    print("\n=== VENCEN DENTRO DE 30 DIAS ===")
    for v, t, h in proximos:
        print(f"  {v}  {h:15s} {t}")
    if not proximos:
        print("  (ninguno)")

    if "--bajas" in argv:
        return 0

    # --- ALTAS: hay que preguntarle a 1816 ---
    from precios_1816 import Cliente1816, Error1816
    cli = Cliente1816()

    # El nombre en el archivo casi nunca es el nombre en 1816: los Bonares y Globales llevan la D
    # que agrega Eco, las ONs cambian la D final por O y los bopreales tienen un mapa propio. Eso
    # ya lo resuelve leer_tickers(), así que se reusa en vez de reimplementarlo. La primera versión
    # comparaba con los nombres del archivo y daba diez altas que ya estaban seguidas: los seis
    # bopreales, BYCVD, T662D, PLC7D y YMCXD.
    conocidos = set(monitor)
    for it in leer_tickers():
        if it["t1816"]:
            conocidos.add(it["t1816"])

    nuevos = {}
    for curva, grupo in CURVAS.items():
        try:
            items = cli.instrumentos(curva_id=curva)
        except Error1816 as e:
            # Una curva que falla no tiene por qué tirar abajo el repaso entero.
            print(f"AVISO: curva {curva} ({grupo}) fallo: {e}", file=sys.stderr)
            continue
        for it in items or []:
            t = str(it.get("ticker") or "").strip()
            if not t or "@" in t or " " in t:      # variantes y opciones
                continue
            if t in conocidos or t in nuevos:
                continue
            venc = a_fecha(it.get("fechaVencimiento"))
            if venc and venc <= hoy:               # ya vencido: no es un alta
                continue
            nuevos[t] = {"grupo": grupo, "venc": venc,
                         "emision": a_fecha(it.get("fechaEmision")),
                         "emisor": it.get("emisorNombre"), "isin": it.get("isinCode"),
                         "moneda": it.get("monedaDenom")}

    falta_sob = {t: v for t, v in nuevos.items() if v["grupo"] in EXHAUSTIVAS}
    falta_otro = {t: v for t, v in nuevos.items() if v["grupo"] not in EXHAUSTIVAS}

    vol = volumenes(cli, sorted(falta_sob))
    print(f"\n=== ALTAS: soberanos que 1816 lista y el monitor no sigue ({len(falta_sob)}) ===")
    for t, v in sorted(falta_sob.items(), key=lambda x: (x[1]["emision"] or date.min)):
        marca = "  <-- liquido" if (vol.get(t) or 0) >= PISO_VOLUMEN else ""
        print(f"  {t:8s} {v['grupo']:12s} emision {v['emision']}  vence {v['venc']}"
              f"  vol {vol.get(t) or 0:>12,.0f}{marca}")
    if not falta_sob:
        print("  (ninguno: el monitor esta al dia)")

    # Provinciales y corporativos: el monitor sigue un subconjunto por liquidez, así que el
    # faltante se mide con volumen antes de proponer nada.
    print(f"\n=== CANDIDATOS provinciales/corporativos no seguidos: {len(falta_otro)} ===")
    if falta_otro:
        vo = volumenes(cli, sorted(falta_otro))
        liquidos = [(v, t) for t, v in vo.items() if v >= PISO_VOLUMEN]
        print(f"  con volumen >= {PISO_VOLUMEN:,}: {len(liquidos)}")
        for vol, t in sorted(liquidos, reverse=True):
            v = falta_otro[t]
            print(f"  {t:8s} {v['grupo']:12s} vol {vol:>12,.0f}  vence {v['venc']}  {v['emisor']}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
