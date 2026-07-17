#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
precios_1816.py — Cliente para levantar precios de mercado desde la API de 1816.

Autentica con tu API Key, obtiene un token temporal (cacheado 24 h), y consulta
el endpoint /v1/mercado/indicadores para uno o más tickers.

------------------------------------------------------------------------------
USO RÁPIDO (línea de comandos)
------------------------------------------------------------------------------
    # 1) Configurar la API Key (una sola vez por sesión de terminal):
    #    PowerShell:  $env:API_1816_KEY = "sk-1816-..."
    #    CMD:         set API_1816_KEY=sk-1816-...
    #    o crear un archivo .1816_key en esta carpeta con la key adentro.

    # Precios de varios tickers a CSV:
    python precios_1816.py --tickers AL30,GD30,GD35 --campos precioClean,tna,paridad -o precios.csv

    # Tickers desde un archivo (txt con uno por línea, o CSV/XLSX con columna 'ticker'):
    python precios_1816.py --tickers-file tickers.txt --campos precioClean,tna -o precios.xlsx

    # Con opciones de mercado:
    python precios_1816.py --tickers AL30 --campos precioClean,tna --moneda ccl --fuente byma

------------------------------------------------------------------------------
USO COMO MÓDULO (desde otro script)
------------------------------------------------------------------------------
    from precios_1816 import Cliente1816
    c = Cliente1816()                       # toma la key de API_1816_KEY o .1816_key
    filas = c.precios(["AL30", "GD30"], ["precioClean", "tna"])
    # filas -> [{"ticker": "AL30", "precioClean": 71.2, "tna": 0.42}, ...]

Requisitos: pip install requests   (openpyxl solo si exportás a .xlsx)
"""

import argparse
import csv
import json
import os
import sys
import time
from datetime import date, timedelta
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Falta la librería 'requests'. Instalala con: pip install requests")


# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------
BASE_URL = "https://api.1816.com.ar"
MODULE = "mercado"

# Campos válidos según la especificación OpenAPI de 1816.
CAMPOS_VALIDOS = {
    "convencionTna", "currentYield", "denominacion", "duration", "durationMod",
    "fechaLiquidacion", "fechaOperacion", "fuente", "moneda", "paridad",
    "plazo", "precioClean", "precioDirty", "spread", "tea", "tem", "ticker",
    "tna", "ultimaOperacion", "volumenMontoDiario", "volumenNominalDiario",
}

# Campos válidos para SERIES históricas (/v1/mercado/series). Difieren de los de
# indicadores: acá solo van campos que son series temporales (más 'valorTecnico').
CAMPOS_SERIES_VALIDOS = {
    "currentYield", "duration", "durationMod", "paridad", "precioClean",
    "precioDirty", "spread", "tea", "tem", "tna", "valorTecnico",
    "volumenMontoDiario", "volumenNominalDiario",
}

MAX_TICKERS_POR_REQUEST = 50      # límite de la API
MIN_SEGUNDOS_ENTRE_REQUESTS = 1.0  # plan Base: máx 1 request/seg

# Ubicación del cache de token y del archivo de key (misma carpeta que el script).
_DIR = Path(__file__).resolve().parent
TOKEN_CACHE = _DIR / ".1816_token.json"
KEY_FILE = _DIR / ".1816_key"


class Error1816(Exception):
    """Error devuelto por la API o el cliente."""


class Cliente1816:
    """Cliente para la API de mercado de 1816."""

    def __init__(self, api_key=None, base_url=BASE_URL):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key or self._cargar_api_key()
        if not self.api_key:
            raise Error1816(
                "No se encontró la API Key. Definí la variable de entorno "
                "API_1816_KEY, o creá un archivo .1816_key, o pasá --api-key."
            )
        self.session = requests.Session()
        self._token = None
        self._ultimo_request = 0.0

    # -- API Key --------------------------------------------------------------
    @staticmethod
    def _cargar_api_key():
        key = os.environ.get("API_1816_KEY")
        if key:
            return key.strip()
        if KEY_FILE.exists():
            return KEY_FILE.read_text(encoding="utf-8").strip()
        return None

    # -- Token ----------------------------------------------------------------
    def _token_valido(self):
        """Devuelve un token válido, usando cache en disco si no expiró."""
        if self._token:
            return self._token

        # Intentar leer del cache (válido si le quedan > 5 min de vida).
        if TOKEN_CACHE.exists():
            try:
                data = json.loads(TOKEN_CACHE.read_text(encoding="utf-8"))
                if data.get("expira_en", 0) - time.time() > 300 and data.get("token"):
                    self._token = data["token"]
                    return self._token
            except (json.JSONDecodeError, OSError):
                pass

        # Pedir un token nuevo.
        self._token = self._obtener_token()
        return self._token

    def _obtener_token(self):
        url = f"{self.base_url}/v1/auth/token"
        try:
            resp = self.session.post(
                url, json={"apiKey": self.api_key, "module": MODULE}, timeout=30
            )
        except requests.RequestException as e:
            raise Error1816(f"No se pudo conectar a {url}: {e}")

        if resp.status_code == 401:
            raise Error1816("API Key inválida (HTTP 401). Verificá la key.")
        if resp.status_code == 403:
            raise Error1816("API Key sin permiso para el módulo 'mercado' (HTTP 403).")
        if not resp.ok:
            raise Error1816(f"Error obteniendo token (HTTP {resp.status_code}): {resp.text}")

        data = resp.json()
        token = data["token"]
        expira_en = time.time() + float(data.get("expiresIn", 86400))
        try:
            TOKEN_CACHE.write_text(
                json.dumps({"token": token, "expira_en": expira_en}), encoding="utf-8"
            )
        except OSError:
            pass  # si no se puede cachear, seguimos igual
        return token

    # -- Throttling ------------------------------------------------------------
    def _esperar_rate_limit(self):
        transcurrido = time.time() - self._ultimo_request
        if transcurrido < MIN_SEGUNDOS_ENTRE_REQUESTS:
            time.sleep(MIN_SEGUNDOS_ENTRE_REQUESTS - transcurrido)
        self._ultimo_request = time.time()

    # -- Request genérico ------------------------------------------------------
    def _get(self, path, params):
        url = f"{self.base_url}{path}"
        for intento in range(2):  # 1 reintento si el token expiró (401)
            self._esperar_rate_limit()
            headers = {"Authorization": f"Bearer {self._token_valido()}"}
            try:
                resp = self.session.get(url, params=params, headers=headers, timeout=60)
            except requests.RequestException as e:
                raise Error1816(f"Fallo de red en {url}: {e}")

            if resp.status_code == 401 and intento == 0:
                # Token vencido/inválido: forzar renovación y reintentar una vez.
                self._token = None
                TOKEN_CACHE.unlink(missing_ok=True)
                continue
            if resp.status_code == 402:
                raise Error1816("Créditos insuficientes (HTTP 402). Revisá tu balance.")
            if resp.status_code == 429:
                raise Error1816("Demasiadas peticiones (HTTP 429). Bajá el ritmo.")
            if not resp.ok:
                raise Error1816(f"Error en {path} (HTTP {resp.status_code}): {resp.text}")
            return resp.json()
        raise Error1816("No se pudo autenticar tras renovar el token.")

    # -- Endpoints -------------------------------------------------------------
    def precios(self, tickers, campos, fuente=None, plazo=None,
                moneda=None, fecha_operacion=None):
        """
        Devuelve una lista de dicts, uno por ticker, con los campos pedidos.

        tickers: lista de strings (se procesa en lotes de 50 automáticamente).
        campos:  lista de strings (ver CAMPOS_VALIDOS).
        """
        tickers = [t.strip().upper() for t in tickers if t and t.strip()]
        if not tickers:
            raise Error1816("No se pasaron tickers.")
        if not campos:
            raise Error1816("No se pasaron campos.")

        invalidos = [c for c in campos if c not in CAMPOS_VALIDOS]
        if invalidos:
            raise Error1816(
                f"Campos inválidos: {', '.join(invalidos)}.\n"
                f"Válidos: {', '.join(sorted(CAMPOS_VALIDOS))}"
            )

        # Aseguramos que 'ticker' venga en la respuesta para poder mapear.
        campos_pedidos = list(dict.fromkeys(campos))  # dedup preservando orden
        campos_req = campos_pedidos if "ticker" in campos_pedidos else campos_pedidos + ["ticker"]

        filas = []
        meta = {}
        for i in range(0, len(tickers), MAX_TICKERS_POR_REQUEST):
            lote = tickers[i:i + MAX_TICKERS_POR_REQUEST]
            params = {"tickers": lote, "campos": campos_req}
            if fuente:
                params["fuente"] = fuente
            if plazo is not None:
                params["plazo"] = plazo
            if moneda:
                params["moneda"] = moneda
            if fecha_operacion:
                params["fechaOperacion"] = fecha_operacion

            data = self._get("/v1/mercado/indicadores", params)
            meta = {k: data.get(k) for k in ("fechaOperacion", "fuente", "plazo", "moneda")}

            instrumentos = data.get("instrumentos", {}) or {}
            for tk in lote:
                datos = instrumentos.get(tk)
                fila = {"ticker": tk}
                if datos is None:
                    fila["_error"] = "sin datos"
                else:
                    for c in campos_pedidos:
                        fila[c] = datos.get(c)
                filas.append(fila)

        self.ultima_meta = meta  # accesible para quien quiera fecha/fuente/moneda
        return filas

    # La API limita cada request de series a un rango máximo de 1 año. Partimos en
    # ventanas de este tamaño (en días) y las concatenamos de forma transparente.
    MAX_DIAS_POR_SERIE = 360

    @staticmethod
    def _ventanas_fechas(fecha_inicial, fecha_final, max_dias=MAX_DIAS_POR_SERIE):
        """Parte [fecha_inicial, fecha_final] en tramos consecutivos de <= max_dias."""
        ini = date.fromisoformat(fecha_inicial)
        fin = date.fromisoformat(fecha_final)
        if ini > fin:
            ini, fin = fin, ini
        ventanas = []
        cur = ini
        while cur <= fin:
            corte = min(cur + timedelta(days=max_dias), fin)
            ventanas.append((cur.isoformat(), corte.isoformat()))
            cur = corte + timedelta(days=1)  # +1 día evita solapar tramos
        return ventanas

    def series(self, tickers, campos, fuente=None, plazo=None, moneda=None,
               fecha_inicial=None, fecha_final=None, convencion_tna=None):
        """
        Devuelve series históricas en formato "tidy": una fila por (ticker, fecha).

        tickers: lista de strings (se procesa en lotes de 50 automáticamente).
        campos:  lista de strings (ver CAMPOS_SERIES_VALIDOS).
        fecha_inicial / fecha_final: 'AAAA-MM-DD'. Si el rango supera 1 año, se parte
            automáticamente en varias requests (la API no acepta >1 año por llamada).
            Si se omiten, la API usa su default (~últimos 30 días).

        Retorno: [{"ticker": "AL30", "fecha": "2026-07-14", "precioClean": 132.2, ...}, ...]
        ordenado por ticker y luego por fecha. La metadata de la corrida queda en
        cliente.ultima_meta (fechaInicial, fechaFinal, fuente, moneda, plazo, convencionTna).
        """
        tickers = [t.strip().upper() for t in tickers if t and t.strip()]
        if not tickers:
            raise Error1816("No se pasaron tickers.")
        if not campos:
            raise Error1816("No se pasaron campos.")

        invalidos = [c for c in campos if c not in CAMPOS_SERIES_VALIDOS]
        if invalidos:
            raise Error1816(
                f"Campos inválidos para series: {', '.join(invalidos)}.\n"
                f"Válidos: {', '.join(sorted(CAMPOS_SERIES_VALIDOS))}"
            )

        campos_pedidos = list(dict.fromkeys(campos))  # dedup preservando orden

        # Ventanas de fechas (auto-chunking del rango si supera 1 año).
        if fecha_inicial and fecha_final:
            ventanas = self._ventanas_fechas(fecha_inicial, fecha_final)
        else:
            ventanas = [(fecha_inicial, fecha_final)]  # single (posible default de la API)

        # tk -> {fecha: {campo: valor}}   (acumula across ventanas)
        acumulado = {tk: {} for tk in tickers}
        meta = {}
        for fi, ff in ventanas:
            meta = self._series_ventana(tickers, campos_pedidos, fuente, plazo,
                                        moneda, fi, ff, convencion_tna, acumulado)

        # La metadata refleja el rango total pedido (no el del último tramo).
        if fecha_inicial:
            meta["fechaInicial"] = fecha_inicial
        if fecha_final:
            meta["fechaFinal"] = fecha_final

        filas = []
        for tk in tickers:
            por_fecha = acumulado[tk]
            if not por_fecha:
                filas.append({"ticker": tk, "fecha": None, "_error": "sin datos"})
                continue
            for fecha in sorted(por_fecha):
                fila = {"ticker": tk, "fecha": fecha}
                for campo in campos_pedidos:
                    fila[campo] = por_fecha[fecha].get(campo)
                filas.append(fila)

        self.ultima_meta = meta
        return filas

    def _series_ventana(self, tickers, campos_pedidos, fuente, plazo, moneda,
                        fecha_inicial, fecha_final, convencion_tna, acumulado):
        """Trae UNA ventana (<=1 año) y vuelca los puntos en `acumulado`. Devuelve meta."""
        meta = {}
        for i in range(0, len(tickers), MAX_TICKERS_POR_REQUEST):
            lote = tickers[i:i + MAX_TICKERS_POR_REQUEST]
            params = {"tickers": lote, "campos": campos_pedidos}
            if fuente:
                params["fuente"] = fuente
            if plazo is not None:
                params["plazo"] = plazo
            if moneda:
                params["moneda"] = moneda
            if fecha_inicial:
                params["fechaInicial"] = fecha_inicial
            if fecha_final:
                params["fechaFinal"] = fecha_final
            if convencion_tna:
                params["convencionTna"] = convencion_tna

            data = self._get("/v1/mercado/series", params)
            meta = {k: data.get(k) for k in (
                "fechaInicial", "fechaFinal", "fuente", "plazo", "moneda", "convencionTna")}

            instrumentos = data.get("instrumentos", {}) or {}
            for tk in lote:
                campos_map = instrumentos.get(tk)
                if not campos_map:
                    continue
                # Pivotar {campo: [[fecha, valor], ...]} -> {fecha: {campo: valor}}
                for campo in campos_pedidos:
                    for punto in (campos_map.get(campo) or []):
                        if not punto:
                            continue
                        fecha = punto[0]
                        valor = punto[1] if len(punto) > 1 else None
                        acumulado[tk].setdefault(fecha, {})[campo] = valor
        return meta

    def balance_creditos(self):
        """Devuelve el balance de créditos del usuario."""
        return self._get("/v1/creditos/balance", {})

    def instrumentos(self, texto=None, solo_performing=False):
        """Lista instrumentos disponibles (opcionalmente filtrados por texto)."""
        params = {"soloPerforming": str(solo_performing).lower()}
        if texto:
            params["texto"] = texto
        return self._get("/v1/mercado/instrumentos", params)


# ---------------------------------------------------------------------------
# Utilidades de entrada / salida
# ---------------------------------------------------------------------------
def leer_tickers_de_archivo(ruta):
    """Lee tickers de un .txt (uno por línea), .csv o .xlsx (columna 'ticker')."""
    p = Path(ruta)
    if not p.exists():
        raise Error1816(f"No existe el archivo: {ruta}")
    ext = p.suffix.lower()

    if ext == ".txt":
        return [ln.strip() for ln in p.read_text(encoding="utf-8").splitlines() if ln.strip()]

    if ext == ".csv":
        with p.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            col = _buscar_columna_ticker(reader.fieldnames)
            return [row[col].strip() for row in reader if row.get(col, "").strip()]

    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise Error1816("Para leer .xlsx instalá openpyxl: pip install openpyxl")
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        filas = ws.iter_rows(values_only=True)
        encabezados = [str(h).strip() if h is not None else "" for h in next(filas)]
        col = _buscar_columna_ticker(encabezados)
        idx = encabezados.index(col)
        tickers = []
        for fila in filas:
            val = fila[idx] if idx < len(fila) else None
            if val is not None and str(val).strip():
                tickers.append(str(val).strip())
        return tickers

    raise Error1816(f"Extensión no soportada: {ext}. Usá .txt, .csv o .xlsx")


def _buscar_columna_ticker(columnas):
    if not columnas:
        raise Error1816("El archivo no tiene encabezados.")
    for c in columnas:
        if c and c.strip().lower() == "ticker":
            return c
    # si no hay columna 'ticker', usar la primera
    return columnas[0]


def escribir_salida(filas, campos_orden, ruta_salida):
    """Escribe las filas a CSV, XLSX o JSON según la extensión de ruta_salida."""
    if not ruta_salida:
        # Imprimir a stdout como tabla simple.
        _imprimir_tabla(filas, campos_orden)
        return

    ext = Path(ruta_salida).suffix.lower()
    if ext == ".json":
        Path(ruta_salida).write_text(
            json.dumps(filas, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    elif ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import Workbook
        except ImportError:
            raise Error1816("Para exportar a .xlsx instalá openpyxl: pip install openpyxl")
        wb = Workbook()
        ws = wb.active
        ws.title = "Precios"
        ws.append(campos_orden)
        for fila in filas:
            ws.append([fila.get(c) for c in campos_orden])
        wb.save(ruta_salida)
    else:  # CSV por defecto
        with open(ruta_salida, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=campos_orden, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(filas)
    print(f"OK: {len(filas)} filas escritas en {ruta_salida}")


def _imprimir_tabla(filas, campos_orden):
    anchos = {c: max(len(c), *(len(str(f.get(c, ""))) for f in filas)) for c in campos_orden} \
        if filas else {c: len(c) for c in campos_orden}
    print(" | ".join(c.ljust(anchos[c]) for c in campos_orden))
    print("-+-".join("-" * anchos[c] for c in campos_orden))
    for fila in filas:
        print(" | ".join(str(fila.get(c, "")).ljust(anchos[c]) for c in campos_orden))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Levanta precios de mercado desde la API de 1816.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    g = parser.add_mutually_exclusive_group(required=True)
    g.add_argument("--tickers", help="Lista de tickers separados por coma. Ej: AL30,GD30")
    g.add_argument("--tickers-file", help="Archivo con tickers (.txt/.csv/.xlsx)")
    g.add_argument("--balance", action="store_true", help="Mostrar balance de créditos y salir")

    parser.add_argument(
        "--campos", default="precioClean,precioDirty,tna,paridad",
        help="Campos separados por coma (default: precioClean,precioDirty,tna,paridad)",
    )
    parser.add_argument("--fuente", choices=["byma", "mae", "homo-1816"], help="Fuente (default byma)")
    parser.add_argument("--plazo", type=int, help="Plazo de liquidación en días (default 1)")
    parser.add_argument("--moneda", choices=["ars", "ccl", "mep"], help="Moneda (default ars)")
    parser.add_argument("--fecha", help="Fecha de operación AAAA-MM-DD (default hoy). Solo modo precios.")
    parser.add_argument("--serie", "--series", dest="serie", action="store_true",
                        help="Modo histórico: trae la serie temporal entre --desde y --hasta")
    parser.add_argument("--desde", help="Fecha inicial AAAA-MM-DD (modo serie; default: la API usa ~30 días atrás)")
    parser.add_argument("--hasta", help="Fecha final AAAA-MM-DD (modo serie; default: hoy)")
    parser.add_argument("--convencion", help="Convención TNA (ej: 180-360). Aplica a series/precios.")
    parser.add_argument("-o", "--output", help="Archivo de salida (.csv/.xlsx/.json). Si se omite, imprime en pantalla.")
    parser.add_argument("--api-key", help="API Key (si no se usa API_1816_KEY ni .1816_key)")

    args = parser.parse_args(argv)

    try:
        cliente = Cliente1816(api_key=args.api_key)

        if args.balance:
            print(json.dumps(cliente.balance_creditos(), ensure_ascii=False, indent=2))
            return 0

        if args.tickers_file:
            tickers = leer_tickers_de_archivo(args.tickers_file)
        else:
            tickers = [t.strip() for t in args.tickers.split(",") if t.strip()]

        campos = [c.strip() for c in args.campos.split(",") if c.strip()]

        if args.serie:
            filas = cliente.series(
                tickers, campos,
                fuente=args.fuente, plazo=args.plazo, moneda=args.moneda,
                fecha_inicial=args.desde, fecha_final=args.hasta,
                convencion_tna=args.convencion,
            )
            # Orden de columnas: ticker, fecha, luego los campos pedidos.
            campos_orden = ["ticker", "fecha"] + [c for c in campos if c not in ("ticker", "fecha")]
        else:
            filas = cliente.precios(
                tickers, campos,
                fuente=args.fuente, plazo=args.plazo,
                moneda=args.moneda, fecha_operacion=args.fecha,
            )
            # Orden de columnas: ticker primero, luego los campos pedidos.
            campos_orden = ["ticker"] + [c for c in campos if c != "ticker"]

        if any("_error" in f for f in filas):
            campos_orden.append("_error")

        meta = getattr(cliente, "ultima_meta", {})
        if meta and args.serie:
            print(f"[desde={meta.get('fechaInicial')} hasta={meta.get('fechaFinal')} "
                  f"fuente={meta.get('fuente')} moneda={meta.get('moneda')} plazo={meta.get('plazo')}]")
        elif meta:
            print(f"[fecha={meta.get('fechaOperacion')} fuente={meta.get('fuente')} "
                  f"moneda={meta.get('moneda')} plazo={meta.get('plazo')}]")

        escribir_salida(filas, campos_orden, args.output)
        return 0

    except Error1816 as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
