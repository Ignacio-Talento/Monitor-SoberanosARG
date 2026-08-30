#!/usr/bin/env python3
"""Series macro largas para la solapa Macro: ITCRM, riesgo país, reservas, TAMAR e IPC.

QUÉ HACE. Deja un único `macro_series.json` en la raíz del repo, que la página levanta con un
fetch y dibuja del lado del cliente. No hay endpoint intermedio a propósito: son series públicas,
lentas —la más rápida se mueve una vez por día— y la más larga arranca en 1997, así que pedirlas en
vivo en cada carga sería pagar tres round-trips a tres servidores distintos para dibujar lo mismo.

DE DÓNDE SALE CADA UNA, que es lo menos obvio:

  · ITCRM — del XLSX oficial, https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/
    ITCRMSerie.xlsx. NO ESTÁ EN LA API de estadísticas monetarias: se buscó "multilateral", "cambio
    real" e "itcrm" en las 1.610 series del catálogo v4.0 el 2026-08-30 y no hay ninguna. Las tres
    series de tipo de cambio que sí tiene son el minorista, el mayorista y el contable.

    El archivo trae el ITCRM y once bilaterales; acá se guarda sólo la columna del multilateral.
    Es CALENDARIO, no rueda: tiene sábados y domingos, porque el BCRA diariza los índices de
    precios. Eso importa para los promedios: un promedio sobre este índice pondera los fines de
    semana igual que los días hábiles, cosa que no pasa con las otras tres series.

  · Riesgo país — argentinadatos, que republica el EMBI+. El BCRA no lo publica. Misma fuente que
    ya usa el informe diario.

  · Reservas (serie 1) y TAMAR de bancos privados en TEA (serie 45) — API del BCRA.

  · IPC mensual (serie 27) — API del BCRA. Va crudo, sin anualizar: la TAMAR real se calcula en el
    browser para que el selector de qué inflación usar no obligue a volver a bajar nada.

  · REM — la base histórica del relevamiento, otro XLSX de URL fija. Sirve para UNA cosa: el INDEC
    publica el IPC de un mes a mediados del siguiente, así que las últimas tres o cuatro semanas de
    la tasa real no tienen con qué deflactarse. Repetir el último dato conocido es asumir que la
    inflación no se movió; el REM es lo que el mercado espera para ESE mes. Se guarda la mediana
    del relevamiento más reciente que haya proyectado cada mes.

LA TAMAR ES CORTA Y NO ES UN BUG. Arranca el 2024-10-01 porque el BCRA creó la serie ahí. Su "Max"
son dos años, contra treinta de las otras tres, y la página lo dice en vez de dejar que se lea como
si el resto de la historia se hubiera perdido.

PAGINACIÓN. La API topea en 3.000 filas por request y no avisa: devuelve 3.000 y listo. Con
`limit=60`, que es lo que usa macro_informe.py para el informe del día, treinta años de reservas
se convierten en dos meses. Por eso acá se pagina con offset hasta que una página vuelva corta.
"""
import json
import sys
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from macro_informe import BCRA, RIESGO_PAIS, UA, _get        # noqa: E402

ITCRM_XLSX = ("https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/ITCRMSerie.xlsx")
# Base histórica del REM. La URL es FIJA y siempre apunta al último relevamiento; la otra planilla
# que publica la página —"tablas-...-jul-2026.xlsx"— lleva el mes en el nombre y habría que
# adivinarlo cada mes.
REM_XLSX = ("https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/"
            "historico-relevamiento-expectativas-mercado.xlsx")
REM_VAR = "Precios minoristas (IPC nivel general; INDEC)"
SALIDA = Path(__file__).parent / "macro_series.json"
DESDE = "1990-01-01"
PAGINA = 3000          # tope real de la API; con más devuelve 3.000 igual


def serie_larga(id_var, desde=DESDE, hasta=None):
    """Serie completa del BCRA, paginando hasta que una página vuelva corta."""
    hasta = hasta or date.today().isoformat()
    filas, off = [], 0
    while True:
        d, _ = _get(f"{BCRA}/{id_var}",
                    {"desde": desde, "hasta": hasta, "limit": PAGINA, "offset": off})
        n = 0
        for bloque in d.get("results", []):
            for x in bloque.get("detalle", []):
                filas.append((x["fecha"], float(x["valor"])))
                n += 1
        if n < PAGINA:
            break
        off += PAGINA
    filas.sort()
    return filas


def riesgo_pais():
    d, _ = _get(RIESGO_PAIS)
    filas = [(x["fecha"], float(x["valor"])) for x in d if x.get("valor") is not None]
    filas.sort()
    return filas


def itcrm():
    """Columna ITCRM del XLSX del BCRA, diaria y de calendario desde 1997."""
    r = requests.get(ITCRM_XLSX, headers={"User-Agent": UA["User-Agent"]}, timeout=180)
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = []
    for f, v in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        # La hoja termina con filas vacías y arranca con dos de encabezado; se filtra por tipo en
        # vez de por número de fila para no depender de que el BCRA no mueva el título.
        if not isinstance(f, datetime) or not isinstance(v, (int, float)):
            continue
        filas.append((f.date().isoformat(), float(v)))
    filas.sort()
    return filas


def rem_ipc():
    """Pronóstico de inflación mensual del REM, uno por mes: la mediana del relevamiento MÁS
    RECIENTE que haya proyectado ese mes.

    PARA QUÉ. El INDEC publica el IPC de un mes a mediados del siguiente, así que la tasa real de
    las últimas tres o cuatro semanas no tiene con qué deflactarse. Repetir el último dato conocido
    es asumir que la inflación no se movió; el REM es el número que el mercado efectivamente
    espera para ESE mes, que es la pregunta que se está haciendo.

    Se queda con la mediana y no con el promedio: el REM tiene entre 30 y 40 participantes y un
    outlier mueve el promedio bastante más que la mediana.

    OJO CON LA COLUMNA "Referencia": la misma variable aparece con "var. % mensual" y con una
    docena de "var. % i.a.; dic-XX". Sin ese filtro se mezclan mensuales con interanuales.
    """
    r = requests.get(REM_XLSX, headers={"User-Agent": UA["User-Agent"]}, timeout=180)
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    ws = wb["Base de Datos Completa"]
    mejor = {}
    for fp, var, ref, per, med, _prom in ws.iter_rows(min_row=3, max_col=6, values_only=True):
        if var != REM_VAR or str(ref).strip() != "var. % mensual":
            continue
        if not isinstance(fp, datetime) or not isinstance(per, datetime) or med is None:
            continue
        k = per.date().isoformat()
        if k not in mejor or fp > mejor[k][0]:
            mejor[k] = (fp, float(med))
    return [(k, mejor[k][1], mejor[k][0].date().isoformat()) for k in sorted(mejor)]


def empaquetar(filas, dec):
    """Dos arrays paralelos en vez de una lista de objetos: pesa un tercio y se dibuja igual."""
    return {"f": [f for f, _ in filas], "v": [round(v, dec) for _, v in filas]}


def main():
    hoy = date.today()
    out = {
        "generado": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "series": {},
        "fallos": [],
    }
    tareas = [
        ("itcrm", "Índice de Tipo de Cambio Real Multilateral", "17-12-15 = 100",
         "BCRA · ITCRMSerie.xlsx", 2, itcrm),
        ("riesgoPais", "Riesgo país · EMBI+ Argentina", "puntos básicos",
         "argentinadatos (EMBI+)", 0, riesgo_pais),
        ("reservas", "Reservas internacionales del BCRA", "millones de USD",
         "BCRA · serie 1", 0, lambda: serie_larga(1)),
        ("tamar", "TAMAR de bancos privados", "% TEA",
         "BCRA · serie 45", 4, lambda: serie_larga(45, "2024-01-01")),
        ("ipc", "IPC nivel general · variación mensual", "% mensual",
         "BCRA/INDEC · serie 27", 2, lambda: serie_larga(27)),
    ]
    try:
        filas = rem_ipc()
        if not filas:
            raise ValueError("sin pronósticos mensuales")
        out["rem"] = {
            "nombre": "IPC mensual esperado · REM del BCRA",
            "fuente": "BCRA · histórico del REM",
            "ultimoRelevamiento": max(f[2] for f in filas),
            "f": [f[0] for f in filas],
            "v": [round(f[1], 3) for f in filas],
            "pron": [f[2] for f in filas],
        }
        print(f"{'rem':11} {len(filas):6} meses  {filas[0][0]} .. {filas[-1][0]}  "
              f"relevamiento {out['rem']['ultimoRelevamiento']}")
    except Exception as e:                                        # noqa: BLE001
        out["fallos"].append(f"rem: {e}")
        print(f"{'rem':11} FALLÓ: {e}")

    for clave, nombre, unidad, fuente, dec, fn in tareas:
        try:
            filas = fn()
            if not filas:
                raise ValueError("serie vacía")
            out["series"][clave] = dict(
                nombre=nombre, unidad=unidad, fuente=fuente,
                desde=filas[0][0], hasta=filas[-1][0], n=len(filas),
                **empaquetar(filas, dec))
            print(f"{clave:11} {len(filas):6} pts  {filas[0][0]} .. {filas[-1][0]}  "
                  f"ult={filas[-1][1]}")
        except Exception as e:                                    # noqa: BLE001
            out["fallos"].append(f"{clave}: {e}")
            print(f"{clave:11} FALLÓ: {e}")

    if not out["series"]:
        raise SystemExit("ninguna serie se pudo bajar; no se pisa el JSON anterior")

    # Si una sola falla se conserva lo que ya había para ESA serie: vale más un ITCRM de ayer que
    # un gráfico vacío, y el JSON dice hasta cuándo llega cada una.
    if out["fallos"] and SALIDA.exists():
        viejo = json.loads(SALIDA.read_text(encoding="utf-8"))
        for clave, val in (viejo.get("series") or {}).items():
            out["series"].setdefault(clave, val)
        if "rem" not in out and viejo.get("rem"):
            out["rem"] = viejo["rem"]

    # SI EL DATO NO CAMBIÓ, NO SE TOCA EL ARCHIVO. El job corre dos veces por día y las series
    # se mueven una: sin esta comparación, cada corrida dejaba un commit cuya única línea
    # distinta era la marca de tiempo —unos 700 commits vacíos por año—. Por eso `generado`
    # significa "cuándo cambió el dato" y no "cuándo corrió el job": lo segundo no le sirve a
    # nadie, y lo primero es justo lo que se quiere saber cuando una serie se queda quieta.
    if SALIDA.exists():
        try:
            viejo = json.loads(SALIDA.read_text(encoding="utf-8"))
            sin_fecha = lambda d: {k: v for k, v in d.items() if k != "generado"}  # noqa: E731
            if sin_fecha(out) == sin_fecha(viejo):
                print(f"\nsin cambios: {SALIDA.name} queda como estaba "
                      f"(generado {viejo.get('generado')})")
                return
        except Exception:                                      # noqa: BLE001
            pass                                              # ilegible: se reescribe y listo

    SALIDA.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")),
                      encoding="utf-8")
    print(f"\n{SALIDA.name}: {SALIDA.stat().st_size / 1024:.0f} KB · {hoy}")


if __name__ == "__main__":
    main()
