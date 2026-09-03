#!/usr/bin/env python3
"""Canje CCL/MEP para el informe, con la misma definición que usa la solapa Glob vs Bon.

QUÉ ES. Cuánto más caro sale el dólar cable que el MEP. Es un solo número por rueda, no uno por
bono, y decide cuánto cuesta mover una posición de un segmento al otro.

DOS FUENTES, EN ESTE ORDEN, que son las mismas de la solapa:

  1 · ÍNDICE DÓLAR BYMA. BYMA arma los índices MEP y CCL con una CANASTA de instrumentos, así que
      no depende de que un bono puntual haya operado ni de un print raro en un segmento fino.
      El canje es `ccl / mep - 1`. La serie sale del widget público del histórico de BYMA —el
      mismo que consume functions/api/byma-dolar.js— y arranca en 2024-01-03.

      Ojo con los headers: sin los `Sec-Fetch-*` y el `Referer` del widget, el host contesta 401
      aunque la URL sea correcta. Están explícitos abajo y no conviene "limpiarlos".

  2 · AL30, del respaldo local series_globvsbon.xlsx. El mismo título cotiza en los dos segmentos,
      así que `canje = precio MEP / precio CCL - 1` y el precio en pesos se cancela: no hace falta
      ningún índice de tipo de cambio.

      VA AL30 Y NO OTRO, y va SOLO. Es el más operado del mercado en las dos puntas —1.432 ruedas
      en MEP contra 1.238 de GD30—, así que es el que menos huecos deja. GD30 estuvo un tiempo como
      control cruzado y se sacó: al ser menos líquido metía dientes de sierra propios que se leían
      como movimientos del canje y no lo eran.

LO QUE NO SE HACE: la MEDIANA del canje implícito en los 40 y pico de bonos que cotizan en las dos
puntas. Se probó y da parecido en el nivel, pero es otra cosa: mezcla la liquidez de cada especie,
así que su dispersión —de 3,0% a 5,7% el 02/09/2026— se lee como si el canje se moviera cuando lo
que cambia es qué bono operó. Para el nivel del canje va un agregado del mercado o un testigo
líquido, no una mediana de instrumentos heterogéneos.

Medido el 12/08/2026, las dos fuentes coinciden en el nivel: índice 3,98% contra 3,91% por AL30.
"""
import json
from datetime import date
from pathlib import Path

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

AJAX = "https://data-widgets.byma.com.ar/wp-admin/admin-ajax.php?action=get_indice_dolar"
REFERER = "https://data-widgets.byma.com.ar/indice-dolar-historico-widget/"
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"),
    "Accept": "*/*",
    "Accept-Language": "es-AR,es;q=0.9",
    "X-Requested-With": "XMLHttpRequest",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "Referer": REFERER,
}
SERIES_XLSX = Path(__file__).parent / "series_globvsbon.xlsx"
BENCH = "AL30"


def _f(v):
    """Los números del widget vienen con coma decimal en algunos campos."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).replace(".", "").replace(",", ".")) if "," in str(v) else float(v)


def desde_byma(timeout=30):
    """-> ({fecha: canje %}, ssl_verificado). Levanta si el widget no responde.

    VERIFICAR EL SSL. El host del widget usa una cadena que Python no valida en Windows —el mismo
    problema que tiene el BCRA—. Si la verificación falla se reintenta sin verificar y se DEJA
    CONSTANCIA en el resultado: es un dato público de sólo lectura y el informe vale más que la
    verificación, pero que quede escrito y no escondido en un except.
    """
    seguro = True
    try:
        r = requests.get(AJAX, headers=HEADERS, timeout=timeout)
    except requests.exceptions.SSLError:
        seguro = False
        r = requests.get(AJAX, headers=HEADERS, timeout=timeout, verify=False)
    r.raise_for_status()
    d = r.json()
    # El widget devuelve {"meta": {...}, "result": [...]} con un OHLC por indice y por rueda.
    # `bymaClosingPrice` es el cierre del indice Dolar BYMA, que es el MEP; `cclClosingPrice`, el
    # del indice CCL. Se usan los CIERRES: el informe habla del cierre de la rueda.
    filas = d.get("result") or []
    out = {}
    for x in filas:
        if not isinstance(x, dict):
            continue
        f = str(x.get("date") or "")[:10]
        mep, ccl = _f(x.get("bymaClosingPrice")), _f(x.get("cclClosingPrice"))
        if len(f) == 10 and mep and ccl:
            out[f] = (ccl / mep - 1) * 100
    if not out:
        raise ValueError(f"el widget respondió pero sin series reconocibles: "
                         f"{json.dumps(d)[:200]}")
    return out, seguro


def desde_al30():
    """-> {fecha: canje %} derivado de AL30 con el respaldo local. No sale a la red."""
    from openpyxl import load_workbook
    wb = load_workbook(SERIES_XLSX, read_only=True, data_only=True)

    def hoja(nombre):
        ws = wb[nombre]
        filas = ws.iter_rows(values_only=True)
        cab = [str(c).strip() if c else "" for c in next(filas)]
        if BENCH not in cab:
            return {}
        i = cab.index(BENCH)
        return {str(r[0])[:10]: r[i] for r in filas
                if r and r[0] and i < len(r) and isinstance(r[i], (int, float))}

    mep, ccl = hoja("Precio MEP"), hoja("Precio CCL")
    return {f: (mep[f] / ccl[f] - 1) * 100 for f in mep if f in ccl and ccl[f]}


def canje(hasta=None, referencias=None):
    """El canje de la rueda `hasta` y sus variaciones, listo para el informe.

    `referencias` es el dict del JSON del informe: {"semanal": "AAAA-MM-DD", ...}. De cada una se
    devuelve el canje de esa rueda —o de la anterior más cercana, avisando— y la diferencia.
    """
    hasta = hasta or date.today().isoformat()
    fuente, serie, aviso, ssl_ok = "índice Dólar BYMA", None, None, True
    try:
        serie, ssl_ok = desde_byma()
        if not ssl_ok:
            aviso = "el certificado del host de BYMA no validó; el dato se trajo sin verificar"
    except Exception as e:                                        # noqa: BLE001
        aviso = f"el índice de BYMA no respondió ({e}); se derivó de {BENCH}"
        fuente, serie = f"{BENCH} en sus dos puntas", desde_al30()

    fechas = sorted(f for f in serie if f <= hasta)
    if not fechas:
        return {"fuente": fuente, "aviso": aviso or "sin datos hasta la fecha pedida"}

    ult = fechas[-1]
    out = {"fuente": fuente, "aviso": aviso, "sslSinVerificar": not ssl_ok,
           "fecha": ult, "valor": serie[ult],
           "exacta": ult == hasta, "ruedas": len(fechas)}
    if len(fechas) > 1:
        out["previo"] = {"fecha": fechas[-2], "valor": serie[fechas[-2]],
                         "variacion": serie[ult] - serie[fechas[-2]]}

    for tipo, ref in (referencias or {}).items():
        antes = [f for f in fechas if f <= ref]
        if antes:
            out[tipo] = {"pedida": ref, "fecha": antes[-1], "valor": serie[antes[-1]],
                         "variacion": serie[ult] - serie[antes[-1]],
                         "exacta": antes[-1] == ref}

    # Rango de lo que va del año, para ubicar el nivel de hoy contra su propia historia.
    delanio = sorted(v for f, v in serie.items() if f[:4] == ult[:4] and f <= hasta)
    if len(delanio) >= 20:
        out["anio"] = {"min": delanio[0], "max": delanio[-1],
                       "mediana": delanio[len(delanio) // 2], "ruedas": len(delanio)}
    return out


if __name__ == "__main__":
    import sys
    hasta = sys.argv[1] if len(sys.argv) > 1 else None
    refs = {"semanal": "2026-08-28", "mensual": "2026-07-31"}
    print(json.dumps(canje(hasta, refs), ensure_ascii=False, indent=2))
