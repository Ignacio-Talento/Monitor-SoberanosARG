#!/usr/bin/env python3
"""
Reexpresa en CCL la historia de los instrumentos que pasaron a valuarse en esa moneda.

POR QUÉ. El monitor pasó a mostrar globales, provinciales y las ONs que pagan en CCL en esa punta.
El histórico seguía en MEP, y eso no es un detalle cosmético: las columnas DAY/WTD/MTD/YTD comparan
el precio de hoy contra un cierre guardado, así que mezclar las dos puntas daba variaciones
infladas por la brecha entre MEP y CCL —números plausibles que no significan lo que dicen—.

QUÉ HACE. Pide a 1816 la serie completa de esos tickers en CCL y PISA los valores que ya están en
historicos.xlsx. Es el único script del repo que sobrescribe: los demás sólo llenan huecos. Por eso
el dry-run es el default y hay que pedir explícitamente que escriba.

CONTROL DE CORDURA. Antes de escribir compara cada valor nuevo contra el que había y reporta la
razón CCL/MEP. Tiene que dar una brecha chica y consistente (del orden de 1,00 a 1,10). Si apareciera
un factor de 1.500 sería que algo vino en pesos, y ahí conviene frenar en vez de escribir.

COMPLETAR CON EL CANJE. El CCL es una punta mucho más fina que el MEP: hay papeles —sobre todo
provinciales y ONs chicas— que tienen precio MEP todos los días y CCL uno de cada tres, o ninguno.
Para esos, pedir la serie en CCL deja huecos, y peor: pisar sólo donde hay dato deja la columna con
las dos monedas mezcladas, que es justo lo que se quería evitar.

El modo --canje los completa convirtiendo el precio MEP por el factor CCL/MEP de esa rueda. El
factor se saca de los instrumentos que SÍ tienen las dos puntas ese día —unos 44, con 0,6% de
dispersión— y coincide con el índice dólar de BYMA dentro del 0,1%, que es una validación
independiente. Convertir no inventa un precio: reexpresa el mismo valor en la otra moneda, que es
exactamente lo que hace 1816 cuando publica las dos.

USO
    python reexpresar_ccl.py                 # dry-run: qué cambiaría y con qué brecha
    python reexpresar_ccl.py --escribir      # aplica
    python reexpresar_ccl.py --canje         # dry-run del completado por canje
    python reexpresar_ccl.py --canje --escribir
    python reexpresar_ccl.py --escribir TLCPD GD30D   # sólo esos
"""

import sys
from statistics import median

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl", file=sys.stderr)
    sys.exit(1)

from actualizar_historicos import CAMPO_1816, HISTORICOS_FILE, leer_tickers

try:
    from precios_1816 import Cliente1816, Error1816, MAX_TICKERS_SERIES
except ImportError:
    print("ERROR: no se pudo importar precios_1816", file=sys.stderr)
    sys.exit(1)

ESPERAS = [15, 45, 90]
# Fuera de esta banda, el valor nuevo no es "el mismo bono en otra punta" sino otra cosa —otra
# moneda, otra unidad—. Se descarta en vez de pisar la serie con basura.
RAZON_MIN, RAZON_MAX = 0.85, 1.30


def leer(ws):
    cab = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    fila_de = {}
    for r in range(2, ws.max_row + 1):
        f = str(ws.cell(row=r, column=1).value or "")[:10]
        if len(f) == 10 and f[4] == "-":
            fila_de[f] = r
    return cab, fila_de


def pedir(cli, tickers, desde, hasta):
    import time
    ultimo = None
    out = []
    for i in range(0, len(tickers), MAX_TICKERS_SERIES):
        lote = tickers[i:i + MAX_TICKERS_SERIES]
        for espera in [0] + ESPERAS:
            if espera:
                print(f"    reintentando en {espera}s...", file=sys.stderr)
                time.sleep(espera)
            try:
                out += cli.series(lote, [CAMPO_1816], moneda="ccl",
                                  fecha_inicial=desde, fecha_final=hasta)
                break
            except Error1816 as e:
                ultimo = e
                if "429" not in str(e) and "Demasiadas" not in str(e):
                    break
            except Exception as e:
                ultimo = e
        else:
            print(f"    lote {i // MAX_TICKERS_SERIES}: sin respuesta ({ultimo})", file=sys.stderr)
    return out


RESPALDO = "historicos_pre_ccl.xlsx"   # la serie tal como estaba, toda en MEP


def completar_con_canje(argv):
    """Rellena en CCL las celdas que quedaron en MEP, usando el canje de cada rueda.

    Se apoya en el respaldo previo a la reexpresión: una celda que hoy vale IGUAL que allá es una
    que 1816 no pudo dar en CCL —el factor es 0,96, así que coincidir a nueve decimales por azar no
    pasa—. Esas son las que se convierten.
    """
    escribir = "--escribir" in argv
    pedidos = {a.upper() for a in argv if not a.startswith("--")}

    enCCL = {it["eco"] for it in leer_tickers() if it["moneda"] == "ccl"
             and (not pedidos or it["eco"] in pedidos)}
    wb = load_workbook(HISTORICOS_FILE)
    ws = wb["Historicos"]
    cab, fila_de = leer(ws)
    col = {t: i + 1 for i, t in enumerate(cab) if t}

    prev = load_workbook(RESPALDO, read_only=True, data_only=True)["Historicos"]
    itp = prev.iter_rows(values_only=True)
    cabp = [str(c) for c in next(itp)]
    mep = {}
    for r in itp:
        if not r or not r[0]:
            continue
        mep[str(r[0])[:10]] = {cabp[i]: r[i] for i in range(1, len(cabp))
                               if isinstance(r[i], (int, float))}

    # Factor por rueda: mediana de (CCL / MEP) sobre los que tienen las dos puntas.
    factor, faltan = {}, []
    for f, fila in mep.items():
        if f not in fila_de:
            continue
        rs = []
        for t, v in fila.items():
            if t not in enCCL or not v or t not in col:
                continue
            act = ws.cell(row=fila_de[f], column=col[t]).value
            if isinstance(act, (int, float)) and abs(act - v) > 1e-9:
                rs.append(act / v)
        if len(rs) >= 3:
            factor[f] = median(rs)
        else:
            faltan.append(f)

    cambios = []
    for f, fila in mep.items():
        if f not in factor:
            continue
        for t, v in fila.items():
            if t not in enCCL or not v or t not in col:
                continue
            celda = ws.cell(row=fila_de[f], column=col[t])
            if isinstance(celda.value, (int, float)) and abs(celda.value - v) <= 1e-9:
                cambios.append((celda, round(v * factor[f], 4), t))

    print(f"{len(factor)} ruedas con canje calculable"
          + (f" · {len(faltan)} sin (menos de 3 pares)" if faltan else ""))
    if factor:
        fs = sorted(factor.values())
        print(f"  canje CCL/MEP: mediana {median(fs):.4f} · mín {fs[0]:.4f} · máx {fs[-1]:.4f}")
    from collections import Counter
    porTicker = Counter(t for _, _, t in cambios)
    print(f"\n{len(cambios)} celdas a completar por canje, en {len(porTicker)} instrumentos")
    for t, n in porTicker.most_common(6):
        print(f"    {t:7s} {n}")
    if not escribir:
        print("\n(dry-run: no se escribió nada. Agregá --escribir para aplicar)")
        return 0
    for celda, v, _ in cambios:
        celda.value = v
    wb.save(HISTORICOS_FILE)
    print(f"\n{HISTORICOS_FILE} actualizado: {len(cambios)} celdas completadas por canje")
    return 0


def main(argv):
    if "--canje" in argv:
        return completar_con_canje(argv)
    escribir = "--escribir" in argv
    pedidos = {a.upper() for a in argv if not a.startswith("--")}

    enCCL = [it for it in leer_tickers()
             if it["moneda"] == "ccl" and it["t1816"]
             and (not pedidos or it["eco"] in pedidos)]
    if not enCCL:
        print("No hay instrumentos en CCL para reexpresar.")
        return 1

    wb = load_workbook(HISTORICOS_FILE)
    ws = wb["Historicos"]
    cab, fila_de = leer(ws)
    col = {t: i + 1 for i, t in enumerate(cab) if t}
    fechas = sorted(fila_de)
    objetivo = [it for it in enCCL if it["eco"] in col]
    print(f"{len(objetivo)} instrumentos en CCL · ruedas {fechas[0]} a {fechas[-1]}")
    if len(objetivo) < len(enCCL):
        print(f"  ({len(enCCL) - len(objetivo)} todavía sin columna en el histórico, se saltean)")

    cli = Cliente1816()
    inv = {it["t1816"]: it["eco"] for it in objetivo}
    filas = pedir(cli, sorted(inv), fechas[0], fechas[-1])

    cambios, razones, raros = [], [], []
    for r in filas:
        v = r.get(CAMPO_1816)
        eco = inv.get(r.get("ticker"))
        f = str(r.get("fecha", ""))[:10]
        if v is None or not eco or f not in fila_de or not isinstance(v, (int, float)) or v <= 0:
            continue
        celda = ws.cell(row=fila_de[f], column=col[eco])
        viejo = celda.value
        if isinstance(viejo, (int, float)) and viejo > 0:
            razon = v / viejo
            if not (RAZON_MIN <= razon <= RAZON_MAX):
                raros.append((eco, f, viejo, v, razon))
                continue
            razones.append(razon)
        cambios.append((celda, v))

    print(f"\n{len(cambios)} celdas a reexpresar")
    if razones:
        razones.sort()
        print(f"  brecha CCL/MEP: mediana {median(razones):.4f} · "
              f"mín {razones[0]:.4f} · máx {razones[-1]:.4f}")
    if raros:
        print(f"  {len(raros)} descartadas por quedar fuera de la banda {RAZON_MIN}-{RAZON_MAX}:")
        for eco, f, viejo, nuevo, razon in raros[:5]:
            print(f"    {eco} {f}: {viejo} -> {nuevo} (x{razon:.2f})")

    if not escribir:
        print("\n(dry-run: no se escribió nada. Agregá --escribir para aplicar)")
        return 0
    for celda, v in cambios:
        celda.value = v
    wb.save(HISTORICOS_FILE)
    print(f"\n{HISTORICOS_FILE} actualizado: {len(cambios)} celdas ahora en CCL")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
